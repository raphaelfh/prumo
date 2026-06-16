"""Legacy sheet writers + public ``build_workbook`` re-export.

The PUBLIC entrypoint ``build_workbook(layout) -> bytes`` is re-exported
from the orchestrator package ``app.services.exports.extraction.workbook``
so the historical import path
``app.services.exports.extraction_xlsx_builder.build_workbook`` keeps
resolving to the *exact same* function object — endpoint/worker imports
stay untouched with zero behaviour drift.

The matrix sheet now lives in the pure sub-builder package
(``app.services.exports.extraction.matrix``). The remaining sheet writers
here (AI metadata + Notes) are still called lazily by the orchestrator
until each migrates to its own pure sub-builder; this module is deleted
once that migration completes.

This module performs NO I/O — no DB session, no storage adapter, no
network. That separation keeps the writers unit-testable without
fixtures.
"""

from __future__ import annotations

from dataclasses import astuple
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.services.exports.value_envelope import format_export_scalar
from app.services.extraction_export_service import (
    AIProposalRow,
    ExportLayout,
)

# ----------------------------------------------------------------------
# Styling — kept small. The reference workbook uses heavier styling;
# spec SC-003 only requires the structural skeleton.
# ----------------------------------------------------------------------

_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


# Re-export the canonical ``build_workbook`` from the orchestrator package
# ``app.services.exports.extraction.workbook`` so the historical import path
# ``app.services.exports.extraction_xlsx_builder.build_workbook`` keeps
# resolving to the *exact same* function object — endpoint/worker imports
# stay untouched with zero behaviour drift. The orchestrator imports the
# sheet writers below lazily, so this top-level re-export introduces no cycle.
from app.services.exports.extraction.workbook import build_workbook  # noqa: E402

# ======================================================================
# AI metadata sheet (T042 — full implementation lands in US1 AI sub-flow)
# ======================================================================


def _write_ai_metadata_sheet(
    workbook: Workbook,
    layout: ExportLayout,
) -> None:
    """Flat-tabular AI metadata sheet (FR-036 – FR-040)."""
    ws = workbook.create_sheet(title="AI metadata")
    headers = [
        "Article",
        "Section",
        "Instance #",
        "Field",
        "AI proposed value",
        "Confidence",
        "Rationale",
        "Evidence text",
        "Evidence page(s)",
        "Proposed at",
        "Reviewer outcome",
        "Final value used",
    ]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    rows: tuple[AIProposalRow, ...] = getattr(layout, "ai_proposal_rows", ()) or ()
    body_row = 2
    # Column indices that carry resolved extraction values (1-based to
    # match the header row); both must render like matrix cells.
    value_col_indices = {5, 12}  # "AI proposed value", "Final value used"
    for row in rows:
        for col_idx, val in enumerate(astuple(row), start=1):
            cell_val = (
                format_export_scalar(val) if col_idx in value_col_indices else _xlsx_safe(val)
            )
            ws.cell(row=body_row, column=col_idx, value=cell_val)
        body_row += 1

    if not rows:
        # FR-039 placeholder line.
        ws.cell(
            row=2,
            column=1,
            value="(No AI proposals recorded for the selected articles.)",
        )


def _xlsx_safe(value: Any) -> Any:
    """Convert values openpyxl cannot serialise natively.

    Lists → joined string; timezone-aware datetimes → naive UTC. A
    ``dict`` here is a bug: ``resolve_value`` is the single unwrapper and
    must have collapsed every envelope upstream. We raise rather than
    silently ``str()`` a Python-repr dict into the workbook (that masked
    the §6 dict-leak in tests).
    """
    if value is None:
        return None
    if isinstance(value, list):
        return "; ".join(str(item) for item in value if item is not None)
    if isinstance(value, dict):
        raise TypeError(
            "_xlsx_safe received a dict; resolve_value must run upstream "
            f"to collapse the envelope (got {value!r})."
        )
    # Datetime handling — openpyxl raises on tz-aware datetimes.
    from datetime import datetime as _dt

    if isinstance(value, _dt) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


__all__ = ["build_workbook"]

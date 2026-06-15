"""Pure intermediate representation for one worksheet + its renderer.

``SheetSpec`` and its value objects are openpyxl-free: every sub-builder
returns a ``SheetSpec`` (or ``list[SheetSpec]`` / ``None``) built from
plain Python, so sub-builder tests assert on rows/cells without a
``Workbook``. ``_render_sheet_spec`` is the ONLY place openpyxl writes to
a worksheet — structural styling only (no conditional formatting, §9).
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CellValue = str | int | float | bool | None


@dataclass(frozen=True)
class CellStyle:
    """Structural-only styling (no conditional formatting — §9)."""

    bold: bool = False
    fill: str | None = None  # hex fill, e.g. "EEEEEE"; None = no fill
    align: str | None = None  # "left" | "center" | "right"
    wrap: bool = False


@dataclass(frozen=True)
class Cell:
    value: CellValue
    style: CellStyle | None = None


@dataclass(frozen=True)
class MergeSpan:
    """1-based inclusive merge range."""

    start_row: int
    start_col: int
    end_row: int
    end_col: int


@dataclass(frozen=True)
class SheetSpec:
    """Pure, openpyxl-free description of one worksheet."""

    title: str  # already sheet-name-safe (<=31, no forbidden chars)
    rows: tuple[tuple[Cell, ...], ...]  # row-major; ragged rows allowed
    merges: tuple[MergeSpan, ...] = ()
    column_widths: tuple[float | None, ...] = ()  # per-column; None = default
    freeze: str | None = None  # openpyxl freeze ref, e.g. "C3"; None = none
    tab_color: str | None = None  # hex tab colour or None


def _style_to_kwargs(
    style: CellStyle,
) -> tuple[Font | None, Alignment | None, PatternFill | None]:
    font = Font(bold=True) if style.bold else None
    alignment = None
    if style.align is not None or style.wrap:
        alignment = Alignment(
            horizontal=style.align,
            vertical="center",
            wrap_text=style.wrap,
        )
    fill = PatternFill("solid", fgColor=style.fill) if style.fill else None
    return font, alignment, fill


def _render_sheet_spec(ws: Worksheet, spec: SheetSpec) -> None:
    """Render a SheetSpec onto an existing (empty) worksheet."""
    ws.title = spec.title

    for r_idx, row in enumerate(spec.rows, start=1):
        for c_idx, cell in enumerate(row, start=1):
            target = ws.cell(row=r_idx, column=c_idx, value=cell.value)
            if cell.style is not None:
                font, alignment, fill = _style_to_kwargs(cell.style)
                if font is not None:
                    target.font = font
                if alignment is not None:
                    target.alignment = alignment
                if fill is not None:
                    target.fill = fill

    for span in spec.merges:
        ws.merge_cells(
            start_row=span.start_row,
            start_column=span.start_col,
            end_row=span.end_row,
            end_column=span.end_col,
        )

    for c_idx, width in enumerate(spec.column_widths, start=1):
        if width is not None:
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    if spec.freeze is not None:
        ws.freeze_panes = spec.freeze

    if spec.tab_color is not None:
        ws.sheet_properties.tabColor = spec.tab_color


__all__ = [
    "Cell",
    "CellStyle",
    "CellValue",
    "MergeSpan",
    "SheetSpec",
    "_render_sheet_spec",
]

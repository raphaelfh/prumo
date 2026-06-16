"""Integration: the resolved ExportLayout renders the new publication sheets.

Real local Supabase (autouse SEED). Scopes all queries by project_id.

End-to-end coverage for Task 54: a Consensus layout resolved for a
finalized run must populate the new service-side projections
(``front_matter`` / ``data_dictionary`` / ``tidy_tables``) and the
orchestrator must render them as the README / Summary / Data dictionary /
per-section tidy-table sheets, carrying baked data.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from unittest.mock import MagicMock
from uuid import UUID

import pytest
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.extraction import ExtractionRun, ExtractionRunStage, TemplateKind
from app.models.extraction_workflow import (
    ExtractionConsensusMode,
    ExtractionProposalSource,
)
from app.services.exports.extraction.workbook import build_workbook
from app.services.extraction_consensus_service import ExtractionConsensusService
from app.services.extraction_export_service import ExportMode, ExtractionExportService
from app.services.extraction_proposal_service import ExtractionProposalService
from app.services.hitl_session_service import HITLSessionService
from app.services.run_lifecycle_service import RunLifecycleService
from tests.integration.conftest import SEED

pytestmark = pytest.mark.asyncio


@dataclass(frozen=True)
class _CharmsExportFixture:
    """Minimal seed handle for the new-sheets export integration test."""

    project_id: UUID
    template_id: UUID
    article_ids: tuple[UUID, ...]
    user_id: UUID
    storage_stub: MagicMock


async def _seed_finalized_charms_run(db: AsyncSession) -> _CharmsExportFixture | None:
    """Drive the seed article through a full extraction run to FINALIZED.

    Mirrors the manual-only / value-resolution flow: open a session,
    park a human proposal, advance PROPOSAL → REVIEW → CONSENSUS,
    publish one value via ``manual_override``, then finalize. Returns
    ``None`` when the dev DB is not seeded so the test can skip.

    All queries are scoped by ``project_id``.
    """
    if (
        await db.execute(
            text("SELECT 1 FROM public.profiles WHERE id = :id"),
            {"id": str(SEED.primary_profile)},
        )
    ).scalar() is None:
        return None

    project_id = SEED.primary_project
    article_id = SEED.primary_article
    template_id = SEED.primary_template
    profile_id = SEED.primary_profile
    instance_id = SEED.primary_instance
    field_id = SEED.primary_field  # seeded 'number' field

    # Fresh run so this test owns the FINALIZED state — the surrounding
    # suite leaks runs in other stages via committed HTTP calls. Scoped
    # by (project_id, article_id, template_id); the closing rollback
    # undoes the cleanup along with the rest of our writes.
    await db.execute(
        text(
            "DELETE FROM public.extraction_runs WHERE project_id = :pid "
            "AND article_id = :aid AND template_id = :tid"
        ),
        {"pid": str(project_id), "aid": str(article_id), "tid": str(template_id)},
    )

    session = await HITLSessionService(db).open_or_resume(
        kind=TemplateKind.EXTRACTION,
        project_id=project_id,
        article_id=article_id,
        user_id=profile_id,
        project_template_id=template_id,
    )
    run_id = session.run_id

    # Park a human proposal so PROPOSAL → REVIEW materializes an
    # accept_proposal reviewer_decision (invariant I-1); the run can
    # then reach CONSENSUS where we publish via manual_override.
    await ExtractionProposalService(db).record_proposal(
        run_id=run_id,
        instance_id=instance_id,
        field_id=field_id,
        source=ExtractionProposalSource.HUMAN,
        source_user_id=profile_id,
        proposed_value={"value": "seed"},
    )

    lifecycle = RunLifecycleService(db)
    await lifecycle.advance_stage(
        run_id=run_id,
        target_stage=ExtractionRunStage.REVIEW,
        user_id=profile_id,
    )
    await lifecycle.advance_stage(
        run_id=run_id,
        target_stage=ExtractionRunStage.CONSENSUS,
        user_id=profile_id,
    )

    _consensus_record, published = await ExtractionConsensusService(db).record_consensus(
        run_id=run_id,
        instance_id=instance_id,
        field_id=field_id,
        consensus_user_id=profile_id,
        mode=ExtractionConsensusMode.MANUAL_OVERRIDE,
        value={"value": {"value": 42, "unit": "patients"}},
        rationale="new-sheets export fixture",
    )
    assert published.run_id == run_id

    await lifecycle.advance_stage(
        run_id=run_id,
        target_stage=ExtractionRunStage.FINALIZED,
        user_id=profile_id,
    )
    run = await db.get(ExtractionRun, run_id)
    assert run is not None
    assert run.stage == ExtractionRunStage.FINALIZED.value

    return _CharmsExportFixture(
        project_id=project_id,
        template_id=template_id,
        article_ids=(article_id,),
        user_id=profile_id,
        storage_stub=MagicMock(),
    )


async def test_resolved_layout_renders_new_sheets(db_session: AsyncSession) -> None:
    """README, Summary, a tidy table, and Data dictionary render with data."""
    fx = await _seed_finalized_charms_run(db_session)
    if fx is None:
        pytest.skip("Missing fixtures.")

    service = ExtractionExportService(
        db=db_session, user_id=str(fx.user_id), storage=fx.storage_stub
    )
    layout = await service.resolve_layout(
        project_id=fx.project_id,
        template_id=fx.template_id,
        mode=ExportMode.CONSENSUS,
        article_ids=list(fx.article_ids),
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
    )

    # Projections are populated.
    assert layout.front_matter is not None
    assert layout.data_dictionary  # >=1 field
    assert layout.tidy_tables  # >=1 section table

    wb = load_workbook(io.BytesIO(build_workbook(layout)))
    assert "README" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "Data dictionary" in wb.sheetnames
    # At least one tidy-table sheet beyond the always-present ones.
    base = {"README", "Summary", "Data dictionary", layout.template_name[:31]}
    assert any(name not in base and name != "Dropdown lists" for name in wb.sheetnames)

    # Front matter carries the template identity + export mode.
    # ``values_only=True`` yields raw cell values (not ``Cell`` objects),
    # so flatten the values directly.
    readme = wb["README"]
    flat = " ".join(
        str(value)
        for row in readme.iter_rows(values_only=True)
        for value in row
        if value is not None
    )
    assert layout.template_name in flat
    assert "Consensus" in flat

    await db_session.rollback()

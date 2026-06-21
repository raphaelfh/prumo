"""Golden-structure assertions for the publication-ready export against
the seeded CHARMS project. Structure only (sheet set + header rows), so
the test is stable across seed-value churn. Scopes by project_id.

Real local Supabase (autouse SEED). The layout is resolved from a freshly
finalized CHARMS run (mirroring ``test_extraction_export_new_sheets.py``)
and the workbook is built once per assertion; we pin the canonical sheet
inventory and the exact header row of the matrix and Data-dictionary sheets
emitted by ``matrix.py`` / ``data_dictionary.py`` / ``front_matter.py``.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from unittest.mock import MagicMock
from uuid import UUID

import pytest
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.extraction import ExtractionRun, ExtractionRunStage, TemplateKind
from app.models.extraction_workflow import (
    ExtractionConsensusMode,
)
from app.services.exports.extraction.workbook import build_workbook
from app.services.extraction_consensus_service import ExtractionConsensusService
from app.services.extraction_export_service import ExportMode, ExtractionExportService
from app.services.hitl_session_service import HITLSessionService
from app.services.run_lifecycle_service import RunLifecycleService
from tests.integration.conftest import SEED

pytestmark = [pytest.mark.asyncio, pytest.mark.integration]

#: Sheets that are NOT per-section tidy tables / the matrix; used to count the
#: data sheets and to find the matrix sheet by exclusion.
_RESERVED = {"Summary", "Data dictionary", "Dropdown lists", "AI metadata"}


@dataclass(frozen=True)
class _CharmsExportFixture:
    """Minimal seed handle for the golden-structure export test."""

    project_id: UUID
    template_id: UUID
    article_ids: tuple[UUID, ...]
    user_id: UUID
    storage_stub: MagicMock


async def _seed_finalized_charms_run(db: AsyncSession) -> _CharmsExportFixture | None:
    """Drive the seed article through a full extraction run to FINALIZED.

    Mirrors the new-sheets / value-resolution flow: open a session, park a
    human proposal, advance PROPOSAL → REVIEW → CONSENSUS, publish one value
    via ``manual_override``, then finalize. Returns ``None`` when the dev DB
    is not seeded so the test can skip. All queries are scoped by ``project_id``.
    """
    if (
        await db.execute(
            text("SELECT 1 FROM public.profiles WHERE id = :id"),
            {"id": str(SEED.primary_profile)},
        )
    ).scalar() is None:
        return None

    project_id = SEED.primary_project
    article_id = SEED.primary_article
    template_id = SEED.primary_template
    profile_id = SEED.primary_profile
    instance_id = SEED.primary_instance
    field_id = SEED.primary_field  # seeded 'number' field

    # Fresh run so this test owns the FINALIZED state — the surrounding suite
    # leaks runs in other stages via committed HTTP calls. Scoped by
    # (project_id, article_id, template_id); the closing rollback undoes this
    # cleanup along with the rest of our writes.
    await db.execute(
        text(
            "DELETE FROM public.extraction_runs WHERE project_id = :pid "
            "AND article_id = :aid AND template_id = :tid"
        ),
        {"pid": str(project_id), "aid": str(article_id), "tid": str(template_id)},
    )

    session = await HITLSessionService(db).open_or_resume(
        kind=TemplateKind.EXTRACTION,
        project_id=project_id,
        article_id=article_id,
        user_id=profile_id,
        project_template_id=template_id,
    )
    run_id = session.run_id

    # The run is already in EXTRACT (session open parks it there); advance
    # straight to CONSENSUS and publish via manual_override.
    lifecycle = RunLifecycleService(db)
    await lifecycle.advance_stage(
        run_id=run_id,
        target_stage=ExtractionRunStage.CONSENSUS,
        user_id=profile_id,
    )

    _consensus_record, published = await ExtractionConsensusService(db).record_consensus(
        run_id=run_id,
        instance_id=instance_id,
        field_id=field_id,
        consensus_user_id=profile_id,
        mode=ExtractionConsensusMode.MANUAL_OVERRIDE,
        value={"value": {"value": 42, "unit": "patients"}},
        rationale="golden-structure export fixture",
    )
    assert published.run_id == run_id

    await lifecycle.advance_stage(
        run_id=run_id,
        target_stage=ExtractionRunStage.FINALIZED,
        user_id=profile_id,
    )
    run = await db.get(ExtractionRun, run_id)
    assert run is not None
    assert run.stage == ExtractionRunStage.FINALIZED.value

    return _CharmsExportFixture(
        project_id=project_id,
        template_id=template_id,
        article_ids=(article_id,),
        user_id=profile_id,
        storage_stub=MagicMock(),
    )


async def _charms_workbook(db: AsyncSession):
    """Resolve a CONSENSUS layout for the seeded CHARMS project and build it.

    Returns ``None`` when the dev DB is not seeded (so callers skip). The
    AI-metadata toggle is ON so the golden inventory pins it as the last sheet.
    """
    fx = await _seed_finalized_charms_run(db)
    if fx is None:
        return None
    service = ExtractionExportService(db=db, user_id=str(fx.user_id), storage=fx.storage_stub)
    layout = await service.resolve_layout(
        project_id=fx.project_id,
        template_id=fx.template_id,
        mode=ExportMode.CONSENSUS,
        article_ids=list(fx.article_ids),
        include_ai_metadata=True,
        anonymize_reviewer_names=False,
    )
    return load_workbook(io.BytesIO(build_workbook(layout)))


async def test_workbook_has_the_expected_sheet_inventory(db_session: AsyncSession) -> None:
    """README/Methods first, Summary + Data dictionary present, AI metadata last."""
    wb = await _charms_workbook(db_session)
    if wb is None:
        pytest.skip("Missing fixtures.")
    titles = wb.sheetnames

    # README/Methods is the front-matter sheet (#1); the sub-builder titles it
    # exactly "README".
    assert titles[0] == "README"
    assert "Summary" in titles
    assert "Data dictionary" in titles
    # AI-metadata toggle is on, so it is appended last (after every §4 sheet).
    assert titles[-1] == "AI metadata"
    # At least the matrix + one tidy table beyond the reserved sheets and README.
    data_sheets = [t for t in titles if t not in _RESERVED and t != "README"]
    assert len(data_sheets) >= 2

    await db_session.rollback()


async def test_data_dictionary_header_is_canonical(db_session: AsyncSession) -> None:
    """The Data-dictionary header row is the exact canonical 8-column order."""
    wb = await _charms_workbook(db_session)
    if wb is None:
        pytest.skip("Missing fixtures.")
    ws = wb["Data dictionary"]
    header = [c.value for c in ws[1]]
    assert header[:8] == [
        "Section",
        "Field",
        "Type",
        "Unit",
        "Description",
        "Allowed values",
        "Required",
        "Allow other",
    ]

    await db_session.rollback()


async def test_matrix_label_columns_are_section_and_field(db_session: AsyncSession) -> None:
    """The matrix sheet (named after the template) leads with Section / Field."""
    wb = await _charms_workbook(db_session)
    if wb is None:
        pytest.skip("Missing fixtures.")
    # The matrix sheet is named after the template; find it by exclusion — it is
    # the first non-reserved, non-README sheet (it is rendered before the tidy
    # tables in §4 order).
    matrix = next(ws for ws in wb.worksheets if ws.title not in _RESERVED and ws.title != "README")
    assert matrix.cell(row=1, column=1).value == "Section"
    assert matrix.cell(row=1, column=2).value == "Field"

    await db_session.rollback()

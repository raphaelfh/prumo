"""Integration coverage for the appraisal-summary sheet (§7).

Exercises ``ExtractionExportService.resolve_layout`` end-to-end against
the real local Supabase Postgres (RLS, CHECK constraints, the real
``version.schema_`` snapshot) for a **quality_assessment** template:

* **Emission gate** — a QA template with risk-label SELECT verdict fields
  yields ``layout.appraisal is not None``; the project's plain
  ``extraction`` template under the same fixture yields
  ``layout.appraisal is None`` (the sections then ship as ordinary tidy
  tables).
* **Mode-aware Overall (§7)** —
    * consensus → one final-score ``Overall`` per record (worst-case
      rollup over the published domain verdicts);
    * all_users → consensus ``Overall`` plus one ``Overall`` per reviewer
      (``per_reviewer_overall`` keyed by ``reviewer_id``, in
      ``layout.reviewers`` order), mirroring the matrix reviewer fan-out;
    * single_user → that reviewer's rollup, with no per-reviewer columns.
* **Render-through** — ``build_workbook`` writes the resolved scalar
  verdicts into the ``Appraisal summary`` cells (no envelope-dict leak).

The fixture builds a self-contained two-domain QA project template
(pointing at the seeded global PROBAST), drives a real run to FINALIZED
publishing the two domain verdicts via the HITL services, and adds a
second reviewer's divergent per-coordinate decisions for the all-users /
single-user cases. Every article/template/instance query is scoped by
``project_id`` (test-fixture scoping rule). The ``db_session`` fixture is
SAVEPOINT-isolated, so every write rolls back at teardown.

Setup mirrors the run / instance / proposal / decision flow in
``test_extraction_manual_only_flow.py`` and the export harness in
``test_extraction_export_new_sheets.py``.
"""

from __future__ import annotations

from dataclasses import dataclass
from unittest.mock import MagicMock
from uuid import UUID, uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.extraction import (
    ExtractionCardinality,
    ExtractionEntityRole,
    ExtractionEntityType,
    ExtractionField,
    ExtractionFieldType,
    ExtractionRun,
    ExtractionRunStage,
    TemplateKind,
)
from app.models.extraction_workflow import (
    ExtractionConsensusMode,
    ExtractionProposalSource,
    ExtractionReviewerDecisionType,
)
from app.services.exports.extraction.workbook import build_workbook
from app.services.extraction_consensus_service import ExtractionConsensusService
from app.services.extraction_export_service import ExportMode, ExtractionExportService
from app.services.extraction_proposal_service import ExtractionProposalService
from app.services.extraction_review_service import ExtractionReviewService
from app.services.hitl_session_service import HITLSessionService
from app.services.run_lifecycle_service import RunLifecycleService
from tests.integration.conftest import SEED

pytestmark = pytest.mark.asyncio

# Seeded global PROBAST quality-assessment template — the project QA
# template points at it via ``global_template_id`` so the fixture mirrors
# a real installed checklist rather than a dangling QA template.
_GLOBAL_PROBAST_ID = UUID("00b00000-0000-0000-0000-000000000001")

# Two appraisal domains, each with one risk-label SELECT verdict field.
_RISK_LABELS = [
    {"value": "Low", "label": "Low"},
    {"value": "Unclear", "label": "Unclear"},
    {"value": "High", "label": "High"},
]


@dataclass(frozen=True)
class _QAFixture:
    """Resolved handles for the appraisal export integration fixture."""

    project_id: UUID
    qa_template_id: UUID
    extraction_template_id: UUID
    article_ids: tuple[UUID, ...]
    user_id: UUID
    reviewer_id: UUID
    domain1_label: str
    domain2_label: str
    storage_stub: MagicMock


async def _seed_finalized_qa_run(
    db: AsyncSession,
    *,
    with_second_reviewer: bool = False,
    domain1_consensus: dict | None = None,
    domain2_consensus: dict | None = None,
) -> _QAFixture | None:
    """Build a 2-domain QA template + drive a run to FINALIZED.

    Returns ``None`` when the dev DB is not seeded so the test can skip.

    Steps (all scoped by ``project_id``):
      1. Insert a ``quality_assessment`` project template with two
         ``study_section`` domains (``cardinality='one'``), each holding
         a single risk-label SELECT verdict field.
      2. ``open_or_resume`` the QA session (materialises one instance per
         domain + a run; the lifecycle snapshots the active version).
      3. Park a human verdict proposal per domain, advance
         PROPOSAL → REVIEW → CONSENSUS, publish ``Low`` / ``High`` via
         ``manual_override``, then FINALIZED.
      4. (all-users / single-user) add a second reviewer so the
         per-reviewer Overall *rollups* diverge: r1 (primary) edits Low on
         BOTH domains -> r1 Overall Low; r2 (reviewer) edits High on both
         -> r2 Overall High. This makes the per-reviewer fan-out
         discriminating (the rollups carry different labels, not just the
         domain-1 inputs).
    """
    if (
        await db.execute(
            text("SELECT 1 FROM public.profiles WHERE id = :id"),
            {"id": str(SEED.primary_profile)},
        )
    ).scalar() is None:
        return None

    project_id = SEED.primary_project
    article_id = SEED.primary_article
    profile_id = SEED.primary_profile
    reviewer_id = SEED.reviewer_profile
    extraction_template_id = SEED.primary_template

    # --- 1. QA project template: two domains, one verdict field each. ---
    qa_template_id = uuid4()
    domain1_et = uuid4()
    domain2_et = uuid4()
    domain1_verdict = uuid4()
    domain2_verdict = uuid4()
    domain1_label = "Participants"
    domain2_label = "Analysis"

    await db.execute(
        text(
            """
            INSERT INTO public.project_extraction_templates
                (id, project_id, global_template_id, name, description,
                 framework, version, kind, schema, is_active, created_by)
            VALUES
                (:id, :pid, :gid, :name, :desc, 'CUSTOM', '1.0.0',
                 :kind, '{}'::jsonb, false, :uid)
            """
        ),
        {
            "id": str(qa_template_id),
            "pid": str(project_id),
            "gid": str(_GLOBAL_PROBAST_ID),
            "name": "Appraisal export QA fixture",
            "desc": "Two-domain risk-of-bias template for the §7 export test.",
            "kind": TemplateKind.QUALITY_ASSESSMENT.value,
            "uid": str(profile_id),
        },
    )

    for et_id, label, sort in (
        (domain1_et, domain1_label, 0),
        (domain2_et, domain2_label, 1),
    ):
        db.add(
            _entity_type(
                et_id=et_id,
                template_id=qa_template_id,
                label=label,
                sort_order=sort,
            )
        )
    await db.flush()

    for verdict_id, et_id in (
        (domain1_verdict, domain1_et),
        (domain2_verdict, domain2_et),
    ):
        # Field 0: a SELECT-typed signalling question (NON risk-label set)
        # that precedes the judgment in sort_order — proves verdict
        # selection keys on the risk-label allowed_values, not position.
        db.add(
            ExtractionField(
                id=uuid4(),
                entity_type_id=et_id,
                name="signalling",
                label="Signalling question",
                field_type=ExtractionFieldType.SELECT.value,
                allowed_values=[
                    {"value": "Y", "label": "Y"},
                    {"value": "PY", "label": "PY"},
                    {"value": "N", "label": "N"},
                ],
                sort_order=0,
            )
        )
        # Field 1: the risk-label SELECT verdict — the appraisal field.
        db.add(
            ExtractionField(
                id=verdict_id,
                entity_type_id=et_id,
                name="risk_of_bias",
                label="Risk of bias",
                field_type=ExtractionFieldType.SELECT.value,
                allowed_values=_RISK_LABELS,
                sort_order=1,
            )
        )
    await db.flush()

    # Fresh run owns the FINALIZED state for the QA template (scoped by the
    # QA template id; the surrounding suite never touches it). The closing
    # rollback undoes everything.
    await db.execute(
        text(
            "DELETE FROM public.extraction_runs WHERE project_id = :pid "
            "AND article_id = :aid AND template_id = :tid"
        ),
        {"pid": str(project_id), "aid": str(article_id), "tid": str(qa_template_id)},
    )

    # --- 2. Open the QA session: instances + run + version snapshot. ---
    session = await HITLSessionService(db).open_or_resume(
        kind=TemplateKind.QUALITY_ASSESSMENT,
        project_id=project_id,
        article_id=article_id,
        user_id=profile_id,
        project_template_id=qa_template_id,
    )
    run_id = session.run_id

    instance_by_et = await _instances_by_entity_type(db, qa_template_id, article_id)
    domain1_instance = instance_by_et[domain1_et]
    domain2_instance = instance_by_et[domain2_et]

    # --- 3. Park human verdict proposals, advance, publish, finalize. ---
    proposal_service = ExtractionProposalService(db)
    p1 = await proposal_service.record_proposal(
        run_id=run_id,
        instance_id=domain1_instance,
        field_id=domain1_verdict,
        source=ExtractionProposalSource.HUMAN,
        source_user_id=profile_id,
        proposed_value={"value": "Low"},
    )
    p2 = await proposal_service.record_proposal(
        run_id=run_id,
        instance_id=domain2_instance,
        field_id=domain2_verdict,
        source=ExtractionProposalSource.HUMAN,
        source_user_id=profile_id,
        proposed_value={"value": "High"},
    )

    lifecycle = RunLifecycleService(db)

    review_service = ExtractionReviewService(db)
    # Primary reviewer accepts both verdicts. The run stays in EXTRACT where
    # these per-reviewer decisions are recorded directly (no boundary
    # materialization — it was removed with the proposal/review collapse).
    await review_service.record_decision(
        run_id=run_id,
        instance_id=domain1_instance,
        field_id=domain1_verdict,
        reviewer_id=profile_id,
        decision=ExtractionReviewerDecisionType.ACCEPT_PROPOSAL,
        proposal_record_id=p1.id,
    )
    await review_service.record_decision(
        run_id=run_id,
        instance_id=domain2_instance,
        field_id=domain2_verdict,
        reviewer_id=profile_id,
        decision=ExtractionReviewerDecisionType.ACCEPT_PROPOSAL,
        proposal_record_id=p2.id,
    )

    if with_second_reviewer:
        # Per-reviewer verdicts that roll up to *different* Overall labels,
        # so the per-reviewer fan-out is discriminating end-to-end: r1
        # (primary) edits Low on BOTH domains -> r1 Overall Low; r2
        # (reviewer) edits High on both -> r2 Overall High. (The earlier
        # accept_proposal on domain 2 published High; the later edit
        # re-points r1's ReviewerState to Low.) ``edit`` decisions carry the
        # value directly into the all-users value map. Consensus stays High
        # (published Low/High) regardless of these reviewer-axis values.
        await review_service.record_decision(
            run_id=run_id,
            instance_id=domain1_instance,
            field_id=domain1_verdict,
            reviewer_id=profile_id,
            decision=ExtractionReviewerDecisionType.EDIT,
            value={"value": "Low"},
        )
        await review_service.record_decision(
            run_id=run_id,
            instance_id=domain2_instance,
            field_id=domain2_verdict,
            reviewer_id=profile_id,
            decision=ExtractionReviewerDecisionType.EDIT,
            value={"value": "Low"},
        )
        await review_service.record_decision(
            run_id=run_id,
            instance_id=domain1_instance,
            field_id=domain1_verdict,
            reviewer_id=reviewer_id,
            decision=ExtractionReviewerDecisionType.EDIT,
            value={"value": "High"},
        )
        await review_service.record_decision(
            run_id=run_id,
            instance_id=domain2_instance,
            field_id=domain2_verdict,
            reviewer_id=reviewer_id,
            decision=ExtractionReviewerDecisionType.EDIT,
            value={"value": "High"},
        )

    await lifecycle.advance_stage(
        run_id=run_id,
        target_stage=ExtractionRunStage.CONSENSUS,
        user_id=profile_id,
    )

    consensus_service = ExtractionConsensusService(db)
    # Published consensus value per domain. Defaults to Low/High; a caller may
    # override with a coded-disposition marker envelope to exercise the ADR-0016
    # appraisal exclusion end-to-end (published-state → resolve_value → sheet).
    d1_consensus = domain1_consensus if domain1_consensus is not None else {"value": "Low"}
    d2_consensus = domain2_consensus if domain2_consensus is not None else {"value": "High"}
    for instance_id, field_id, cvalue in (
        (domain1_instance, domain1_verdict, d1_consensus),
        (domain2_instance, domain2_verdict, d2_consensus),
    ):
        _record, published = await consensus_service.record_consensus(
            run_id=run_id,
            instance_id=instance_id,
            field_id=field_id,
            consensus_user_id=profile_id,
            mode=ExtractionConsensusMode.MANUAL_OVERRIDE,
            value=cvalue,
            rationale="appraisal export fixture",
        )
        assert published.run_id == run_id

    await lifecycle.advance_stage(
        run_id=run_id,
        target_stage=ExtractionRunStage.FINALIZED,
        user_id=profile_id,
    )
    run = await db.get(ExtractionRun, run_id)
    assert run is not None
    assert run.stage == ExtractionRunStage.FINALIZED.value

    return _QAFixture(
        project_id=project_id,
        qa_template_id=qa_template_id,
        extraction_template_id=extraction_template_id,
        article_ids=(article_id,),
        user_id=profile_id,
        reviewer_id=reviewer_id,
        domain1_label=domain1_label,
        domain2_label=domain2_label,
        storage_stub=MagicMock(),
    )


def _entity_type(
    *,
    et_id: UUID,
    template_id: UUID,
    label: str,
    sort_order: int,
) -> ExtractionEntityType:
    return ExtractionEntityType(
        id=et_id,
        project_template_id=template_id,
        name=label.lower(),
        label=label,
        role=ExtractionEntityRole.STUDY_SECTION.value,
        cardinality=ExtractionCardinality.ONE.value,
        parent_entity_type_id=None,
        sort_order=sort_order,
        is_required=False,
    )


async def _instances_by_entity_type(
    db: AsyncSession,
    template_id: UUID,
    article_id: UUID,
) -> dict[UUID, UUID]:
    """Map ``entity_type_id -> instance_id`` for the QA run's instances.

    Scoped by ``(template_id, article_id)`` per the fixture-scoping rule.
    """
    rows = (
        await db.execute(
            text(
                "SELECT entity_type_id, id FROM public.extraction_instances "
                "WHERE template_id = :tid AND article_id = :aid"
            ),
            {"tid": str(template_id), "aid": str(article_id)},
        )
    ).all()
    return {UUID(str(et)): UUID(str(iid)) for et, iid in rows}


def _service(db: AsyncSession, fx: _QAFixture) -> ExtractionExportService:
    return ExtractionExportService(db=db, user_id=str(fx.user_id), storage=fx.storage_stub)


async def test_qa_consensus_layout_emits_appraisal_and_extraction_does_not(
    db_session: AsyncSession,
) -> None:
    """Consensus QA layout carries the appraisal model; extraction does not.

    The QA template's two domains roll up to a single worst-case
    ``Overall`` per record (``Low`` ∨ ``High`` → ``High``). The same
    project's plain ``extraction`` template yields ``appraisal is None``
    (emission gate — the sections ship as ordinary tidy tables).
    """
    fx = await _seed_finalized_qa_run(db_session)
    if fx is None:
        pytest.skip("Missing fixtures.")

    layout = await _service(db_session, fx).resolve_layout(
        project_id=fx.project_id,
        template_id=fx.qa_template_id,
        mode=ExportMode.CONSENSUS,
        article_ids=list(fx.article_ids),
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
    )

    assert layout.appraisal is not None
    assert layout.appraisal.domain_labels == (fx.domain1_label, fx.domain2_label)
    assert len(layout.appraisal.rows) == 1
    row = layout.appraisal.rows[0]
    assert row.domain_verdicts == ("Low", "High")
    assert row.overall == "High"
    assert row.per_reviewer_overall == {}

    # Emission gate: the project's extraction template carries no
    # appraisal layer.
    extraction_layout = await _service(db_session, fx).resolve_layout(
        project_id=fx.project_id,
        template_id=fx.extraction_template_id,
        mode=ExportMode.CONSENSUS,
        article_ids=list(fx.article_ids),
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
    )
    assert extraction_layout.appraisal is None

    await db_session.rollback()


async def test_qa_all_users_and_single_user_overall(db_session: AsyncSession) -> None:
    """All-users adds a per-reviewer Overall; single-user is one rollup.

    All-users: consensus ``Overall`` matches the published worst-case
    (``High``) and ``per_reviewer_overall`` carries one entry per reviewer
    whose rollup is that reviewer's *own* worst-case verdict — which here
    diverges by reviewer: r1 (primary) edits Low on both domains → Overall
    ``Low``; r2 (reviewer) edits High on both → Overall ``High``. The
    divergent labels make the fan-out discriminating: a builder that wrongly
    echoed the consensus Overall into every reviewer's entry would map r1 to
    ``High`` and fail here.

    Single-user (reviewer r2): that reviewer's rollup is ``High`` and the
    per-reviewer map is empty (no reviewer-axis fan-out).
    """
    fx = await _seed_finalized_qa_run(db_session, with_second_reviewer=True)
    if fx is None:
        pytest.skip("Missing fixtures.")

    all_users = await _service(db_session, fx).resolve_layout(
        project_id=fx.project_id,
        template_id=fx.qa_template_id,
        mode=ExportMode.ALL_USERS,
        article_ids=list(fx.article_ids),
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
    )
    assert all_users.appraisal is not None
    row = all_users.appraisal.rows[0]
    # Consensus Overall mirrors the published state (Low ∨ High → High),
    # independent of the reviewer-axis values below.
    assert row.overall == "High"
    # One per-reviewer Overall entry per reviewer, in layout.reviewers order.
    reviewer_ids = [r.reviewer_id for r in all_users.reviewers]
    assert set(row.per_reviewer_overall.keys()) == set(reviewer_ids)
    assert reviewer_ids, "all-users must surface ≥1 reviewer with decisions"
    # Discriminating: each reviewer's rollup is their OWN worst-case, so the
    # labels diverge — r1 (primary) Low, r2 (reviewer) High.
    assert row.per_reviewer_overall[fx.user_id] == "Low"
    assert row.per_reviewer_overall[fx.reviewer_id] == "High"

    single_user = await _service(db_session, fx).resolve_layout(
        project_id=fx.project_id,
        template_id=fx.qa_template_id,
        mode=ExportMode.SINGLE_USER,
        article_ids=list(fx.article_ids),
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        reviewer_id=fx.reviewer_id,
    )
    assert single_user.appraisal is not None
    su_row = single_user.appraisal.rows[0]
    assert su_row.overall == "High"
    assert su_row.per_reviewer_overall == {}

    await db_session.rollback()


async def test_qa_appraisal_renders_through_to_workbook(db_session: AsyncSession) -> None:
    """The resolved verdicts reach the ``Appraisal summary`` cells.

    Builds the consensus QA workbook and asserts the sheet exists, its
    header row is ``(Record, <d1>, <d2>, Overall)`` and the data row's
    last cell is the resolved scalar ``"High"`` — proving no envelope dict
    leaks into the appraisal cells (the value map fed already-resolved
    scalars).
    """
    fx = await _seed_finalized_qa_run(db_session)
    if fx is None:
        pytest.skip("Missing fixtures.")

    layout = await _service(db_session, fx).resolve_layout(
        project_id=fx.project_id,
        template_id=fx.qa_template_id,
        mode=ExportMode.CONSENSUS,
        article_ids=list(fx.article_ids),
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
    )

    wb = load_workbook(__import__("io").BytesIO(build_workbook(layout)))
    assert "Appraisal summary" in wb.sheetnames
    sheet = wb["Appraisal summary"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Record", fx.domain1_label, fx.domain2_label, "Overall")
    # One data row; its Overall (last cell) is the resolved scalar.
    assert rows[1][-1] == "High"

    await db_session.rollback()


async def test_qa_appraisal_marker_verdict_excluded_from_overall(
    db_session: AsyncSession,
) -> None:
    """ADR-0016 Phase 4, end-to-end: a coded-disposition verdict published into
    ExtractionPublishedState survives the read + ``resolve_value`` and reaches the
    Appraisal sheet as its stable LABEL (never a dict-stringify), while the
    worst-case Overall EXCLUDES it — so a "the source is silent" verdict cannot
    silently force a most-severe Overall. Pins the DB → resolve → sheet coupling
    the pure unit tests can't (they start from already-resolved scalars).
    """
    fx = await _seed_finalized_qa_run(
        db_session,
        domain1_consensus={"value": None, "absent_reason": "no_information"},
    )
    if fx is None:
        pytest.skip("Missing fixtures.")

    layout = await _service(db_session, fx).resolve_layout(
        project_id=fx.project_id,
        template_id=fx.qa_template_id,
        mode=ExportMode.CONSENSUS,
        article_ids=list(fx.article_ids),
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
    )

    wb = load_workbook(__import__("io").BytesIO(build_workbook(layout)))
    sheet = wb["Appraisal summary"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Record", fx.domain1_label, fx.domain2_label, "Overall")
    # domain1 cell renders the resolved disposition LABEL — not a dict, not blank.
    assert rows[1][1] == "No information"
    assert rows[1][2] == "High"
    # Overall excludes the marker → the real "High" wins (NOT "No information").
    assert rows[1][-1] == "High"

    await db_session.rollback()

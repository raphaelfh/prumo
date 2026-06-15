"""Pure IR (SheetSpec) + the single openpyxl renderer. The IR is
openpyxl-free so every sub-builder is testable without a Workbook; the
renderer is the only place openpyxl touches a worksheet."""

from __future__ import annotations

import dataclasses

from openpyxl import Workbook

from app.services.exports.extraction.sheet_spec import (
    Cell,
    CellStyle,
    MergeSpan,
    SheetSpec,
    _render_sheet_spec,
)


def test_value_objects_are_frozen() -> None:
    for cls in (Cell, CellStyle, MergeSpan, SheetSpec):
        assert dataclasses.is_dataclass(cls)
        assert cls.__dataclass_params__.frozen is True


def test_cell_defaults_to_no_style() -> None:
    assert Cell(value="x").style is None


def test_render_writes_values_and_merges_and_freeze_and_tab() -> None:
    spec = SheetSpec(
        title="Demo",
        rows=(
            (Cell("A1", CellStyle(bold=True, fill="EEEEEE")), Cell("B1")),
            (Cell(5), Cell(2.5)),
        ),
        merges=(MergeSpan(start_row=1, start_col=1, end_row=1, end_col=2),),
        column_widths=(16.0, None),
        freeze="B2",
        tab_color="FF0000",
    )
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="placeholder")

    _render_sheet_spec(ws, spec)

    assert ws.title == "Demo"
    assert ws["A1"].value == "A1"
    assert ws["A1"].font.bold is True
    assert ws["A2"].value == 5
    assert ws["B2"].value == 2.5
    assert "A1:B1" in {str(m) for m in ws.merged_cells.ranges}
    assert ws.freeze_panes == "B2"
    assert ws.sheet_properties.tabColor is not None
    assert ws.column_dimensions["A"].width == 16.0


def test_render_skips_none_cells_and_ragged_rows() -> None:
    spec = SheetSpec(
        title="Ragged",
        rows=(
            (Cell(None), Cell("kept")),
            (Cell("only-one-col"),),
        ),
    )
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="x")
    _render_sheet_spec(ws, spec)
    assert ws["A1"].value is None
    assert ws["B1"].value == "kept"
    assert ws["A2"].value == "only-one-col"

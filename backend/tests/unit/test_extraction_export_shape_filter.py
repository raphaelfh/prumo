"""Output-shape filter over ``_ordered_specs`` (§3).

Three shapes are subsets of the sheets the workbook already builds. The
filter keys on BUILDER IDENTITY, not on sheet title: ``SheetSpec`` carries
only a title, and a title-matching filter would misfire on an instrument
with a section literally named "Data dictionary". These tests pin both
halves — the emitted sheet set, and that a filtered-out builder is never
called (the filter must save the work, not discard it afterwards).
"""

from __future__ import annotations

import io
from dataclasses import replace
from uuid import UUID, uuid4

from openpyxl import load_workbook

from app.models.extraction import (
    ExtractionCardinality,
    ExtractionEntityRole,
    ExtractionFieldType,
)
from app.services.exports.extraction import workbook as workbook_module
from app.services.exports.extraction.sheet_spec import Cell, SheetSpec
from app.services.exports.extraction.workbook import (
    ExportShape,
    build_workbook,
    shape_needs_articles,
)
from app.services.exports.extraction_snapshot_reader import AllowedValue
from app.services.extraction_export_service import (
    AppraisalModel,
    AppraisalRow,
    ArticleDescriptor,
    ExportLayout,
    ExportMode,
    ExportNotes,
    FieldDescriptor,
    FieldDictEntry,
    SectionDescriptor,
    TidyRow,
    TidyTable,
)

_SECTION_ID = UUID("11111111-1111-1111-1111-111111111111")
_FIELD_ID = UUID("22222222-2222-2222-2222-222222222222")
_ARTICLE_ID = UUID("33333333-3333-3333-3333-333333333333")


def _every_sheet_layout() -> ExportLayout:
    """A layout that makes EVERY conditional builder emit.

    Appraisal, dropdown lists and AI metadata are each conditional; a layout
    that skipped any of them would let a shape look correct for the wrong
    reason (an absent sheet rather than a filtered one).
    """
    field = FieldDescriptor(
        field_id=_FIELD_ID,
        label="Risk of bias",
        type=ExtractionFieldType.SELECT,
        allowed_values=("Low", "High"),
    )
    section = SectionDescriptor(
        entity_type_id=_SECTION_ID,
        label="Participants",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
    )
    instance_id = uuid4()
    article = ArticleDescriptor(
        article_id=_ARTICLE_ID,
        header_label="Gaca, 2011",
        run_id=uuid4(),
        version_id=None,
        model_instances=(),
        section_instances={_SECTION_ID: (instance_id,)},
    )
    return ExportLayout(
        project_name="P",
        template_name="PROBAST",
        template_version=1,
        sections=(section,),
        articles=(article,),
        reviewers=(),
        mode=ExportMode.CONSENSUS,
        include_ai_metadata=True,
        anonymize_reviewer_names=False,
        notes=ExportNotes(),
        value_map={(article.run_id, instance_id, _FIELD_ID): "Low"},
        data_dictionary=(
            FieldDictEntry(
                field_id=_FIELD_ID,
                section_label="Participants",
                label="Risk of bias",
                type=ExtractionFieldType.SELECT,
                unit=None,
                description=None,
                allowed_values=(
                    AllowedValue(value="low", label="Low"),
                    AllowedValue(value="high", label="High"),
                ),
                is_required=True,
                allow_other=False,
            ),
        ),
        tidy_tables=(
            TidyTable(
                section_id=_SECTION_ID,
                title="Participants",
                cardinality=ExtractionCardinality.ONE,
                column_field_ids=(_FIELD_ID,),
                column_labels=("Risk of bias",),
                rows=(
                    TidyRow(
                        article_id=_ARTICLE_ID,
                        instance_id=instance_id,
                        record_label="Gaca, 2011",
                        values=("Low",),
                    ),
                ),
            ),
        ),
        appraisal=AppraisalModel(
            domain_section_ids=(_SECTION_ID,),
            domain_labels=("Participants",),
            rows=(
                AppraisalRow(
                    article_id=_ARTICLE_ID,
                    record_label="Gaca, 2011",
                    domain_verdicts=("Low",),
                    overall="Low",
                    per_reviewer_overall={},
                ),
            ),
        ),
    )


def _sheetnames(shape: ExportShape) -> list[str]:
    data = build_workbook(_every_sheet_layout(), shape=shape)
    return load_workbook(io.BytesIO(data)).sheetnames


def test_complete_is_the_default_and_emits_every_sheet() -> None:
    every = [
        "README",
        "Summary",
        "PROBAST",
        "Participants",
        "Appraisal summary",
        "Data dictionary",
        "Dropdown lists",
        "AI metadata",
    ]
    assert _sheetnames(ExportShape.COMPLETE) == every
    # No shape argument = today's behaviour, byte-for-byte the same sheets.
    default = load_workbook(io.BytesIO(build_workbook(_every_sheet_layout())))
    assert default.sheetnames == every


def test_dictionary_keeps_readme_and_the_catalogue_only() -> None:
    assert _sheetnames(ExportShape.DICTIONARY) == [
        "README",
        "Data dictionary",
        "Dropdown lists",
    ]


def test_publication_keeps_readme_the_tidy_tables_and_the_appraisal() -> None:
    assert _sheetnames(ExportShape.PUBLICATION) == [
        "README",
        "Participants",
        "Appraisal summary",
    ]


def test_readme_is_in_every_shape() -> None:
    """It carries the template identity, provenance and glyph legend."""
    for shape in ExportShape:
        assert "README" in _sheetnames(shape)


def _spy_builders(monkeypatch, shape: ExportShape) -> set[str]:
    """Record which sub-builders ``_ordered_specs`` actually invokes."""
    called: set[str] = set()

    def _spy(name: str, result):
        def _fn(_layout):
            called.add(name)
            return result

        return _fn

    one = SheetSpec(title="x", rows=((Cell("x"),),))
    for name, result in (
        ("front_matter", one),
        ("summary", one),
        ("matrix", one),
        ("appraisal_summary", one),
        ("data_dictionary", one),
        ("dropdown_lists", one),
        ("ai_metadata", one),
    ):
        monkeypatch.setattr(workbook_module, f"build_{name}", _spy(name, result))
    monkeypatch.setattr(workbook_module, "build_tidy_tables", _spy("tidy_tables", []))

    workbook_module._ordered_specs(_every_sheet_layout(), shape)
    return called


def test_dictionary_shape_never_calls_the_filtered_out_builders(monkeypatch) -> None:
    called = _spy_builders(monkeypatch, ExportShape.DICTIONARY)
    assert called == {"front_matter", "data_dictionary", "dropdown_lists"}


def test_publication_shape_never_calls_the_filtered_out_builders(monkeypatch) -> None:
    called = _spy_builders(monkeypatch, ExportShape.PUBLICATION)
    assert called == {"front_matter", "tidy_tables", "appraisal_summary"}


def test_complete_shape_calls_every_builder(monkeypatch) -> None:
    called = _spy_builders(monkeypatch, ExportShape.COMPLETE)
    assert called == {
        "front_matter",
        "summary",
        "matrix",
        "tidy_tables",
        "appraisal_summary",
        "data_dictionary",
        "dropdown_lists",
        "ai_metadata",
    }


# ---------------------------------------------------------------------------
# A dictionary-only workbook describes the INSTRUMENT, so it must build before
# anyone has finalized a run — the state a project is in while it is still
# being wired up.
# ---------------------------------------------------------------------------


def _no_article_layout() -> ExportLayout:
    """The template resolved for a project with zero eligible articles."""
    layout = _every_sheet_layout()
    return replace(layout, articles=(), tidy_tables=(), appraisal=None, value_map={})


def test_dictionary_shape_builds_with_no_articles() -> None:
    data = build_workbook(_no_article_layout(), shape=ExportShape.DICTIONARY)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["README", "Data dictionary", "Dropdown lists"]
    # The catalogue is the point: it must carry the field, not just a header.
    assert any(
        "Risk of bias" in [str(c) for c in row]
        for row in wb["Data dictionary"].iter_rows(values_only=True)
    )


def test_only_the_dictionary_shape_is_free_of_article_values() -> None:
    assert shape_needs_articles(ExportShape.COMPLETE) is True
    assert shape_needs_articles(ExportShape.PUBLICATION) is True
    assert shape_needs_articles(ExportShape.DICTIONARY) is False

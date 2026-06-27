"""Unit tests for the extraction XLSX builder.

Pure-function tests that exercise the builder with hand-crafted
ExportLayout instances. No DB or storage involved.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from uuid import UUID, uuid4

import pytest
from openpyxl import load_workbook

from app.models.extraction import ExtractionEntityRole, ExtractionFieldType
from app.services.exports.extraction.workbook import build_workbook
from app.services.extraction_export_service import (
    ArticleDescriptor,
    ExportLayout,
    ExportMode,
    ExportNotes,
    FieldDescriptor,
    SectionDescriptor,
)

# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------


def _field(label: str, ftype: ExtractionFieldType, parent: UUID) -> FieldDescriptor:
    return FieldDescriptor(
        field_id=uuid4(),
        label=label,
        type=ftype,
        allowed_values=(),
        parent_section_id=parent,
    )


def _section(
    label: str,
    role: ExtractionEntityRole,
    fields: list[FieldDescriptor] | None = None,
    parent: UUID | None = None,
) -> SectionDescriptor:
    eid = uuid4()
    # Re-parent the fields to this section id so the descriptor coherence holds.
    f = tuple(
        FieldDescriptor(
            field_id=f.field_id,
            label=f.label,
            type=f.type,
            allowed_values=f.allowed_values,
            parent_section_id=eid,
        )
        for f in (fields or [])
    )
    return SectionDescriptor(
        entity_type_id=eid,
        label=label,
        role=role,
        parent_entity_type_id=parent,
        fields=f,
    )


def _article(
    header: str,
    *,
    study_instances: dict[UUID, UUID],
    model_instances: tuple[UUID, ...] = (),
    run_id: UUID | None = None,
) -> ArticleDescriptor:
    return ArticleDescriptor(
        article_id=uuid4(),
        header_label=header,
        run_id=run_id if run_id is not None else uuid4(),
        run_stage=None,  # not consulted by builder
        version_id=None,
        model_instances=model_instances,
        # ``study_instances`` is now a read-compat alias property; build the
        # ordered ``section_instances`` from the legacy single-id-per-section
        # argument so existing call sites stay unchanged.
        section_instances={sid: (iid,) for sid, iid in study_instances.items()},
    )


def _layout(
    *,
    sections: tuple[SectionDescriptor, ...] = (),
    articles: tuple[ArticleDescriptor, ...] = (),
    value_map: dict | None = None,
    include_ai_metadata: bool = False,
    project_name: str = "Test Project",
    template_name: str = "CHARMS",
    appraisal: object | None = None,
) -> ExportLayout:
    return ExportLayout(
        project_name=project_name,
        template_name=template_name,
        template_version=1,
        sections=sections,
        articles=articles,
        reviewers=(),
        mode=ExportMode.CONSENSUS,
        include_ai_metadata=include_ai_metadata,
        anonymize_reviewer_names=False,
        notes=ExportNotes(
            template_version_label=f"{template_name} v1",
            export_mode_label="Consensus",
            generated_at=datetime(2026, 5, 23, 12, 0, 0, tzinfo=UTC),
        ),
        value_map=value_map or {},
        appraisal=appraisal,
    )


def _open(data: bytes):
    return load_workbook(io.BytesIO(data))


# ----------------------------------------------------------------------
# Smoke tests (foundation phase, kept for regression)
# ----------------------------------------------------------------------


def test_build_workbook_returns_valid_xlsx_bytes():
    data = build_workbook(_layout())
    assert isinstance(data, bytes)
    assert data[:4] == b"PK\x03\x04"
    wb = _open(data)
    # §4 order: README (absorbs the old Notes sheet) → Summary → matrix →
    # Data dictionary. The empty layout has no sections (no tidy tables) and
    # an empty data dictionary (no Dropdown lists sheet).
    assert wb.sheetnames == ["README", "Summary", "CHARMS", "Data dictionary"]


def test_build_workbook_includes_ai_metadata_sheet_when_toggled():
    data = build_workbook(_layout(include_ai_metadata=True))
    wb = _open(data)
    # AI metadata is the trailing optional sheet, appended after the §4 specs.
    assert wb.sheetnames == [
        "README",
        "Summary",
        "CHARMS",
        "Data dictionary",
        "AI metadata",
    ]


def test_sheet_name_is_sanitised_for_openpyxl_constraints():
    data = build_workbook(_layout(template_name="Bad/Name?:With*Forbidden[chars]"))
    wb = _open(data)
    # The matrix sheet (named from the template) is the only one derived from
    # the unsafe template name; the fixed sheets (README/Summary/...) are safe.
    main = next(s for s in wb.sheetnames if s.startswith("BadName"))
    assert len(main) <= 31
    for forbidden in r"[]:*?/\\":
        assert forbidden not in main


# ----------------------------------------------------------------------
# Consensus layout — single study section, one article (T027 (a))
# ----------------------------------------------------------------------


def test_single_article_single_section_single_field_consensus():
    # The builder now owns hierarchical numbering (§9), so fixtures carry the
    # bare labels and we assert on the builder-generated "1." / "1.1" prefixes.
    f = _field("Source of data", ExtractionFieldType.TEXT, parent=uuid4())
    section = _section("Source of data", ExtractionEntityRole.STUDY_SECTION, [f])
    inst_id = uuid4()
    article = _article("Gaca, 2011", study_instances={section.entity_type_id: inst_id})
    field_id = section.fields[0].field_id

    data = build_workbook(
        _layout(
            sections=(section,),
            articles=(article,),
            value_map={(article.run_id, inst_id, field_id): "Existing registry"},
        )
    )
    ws = _open(data)["CHARMS"]

    # Row 1 = headers; row 2 = section name; row 3 = field row
    assert ws.cell(row=1, column=1).value == "Section"
    assert ws.cell(row=1, column=2).value == "Field"
    assert ws.cell(row=1, column=3).value == "Gaca, 2011"
    assert ws.cell(row=2, column=1).value == "1. Source of data"
    assert ws.cell(row=3, column=2).value == "1.1 Source of data"
    assert ws.cell(row=3, column=3).value == "Existing registry"


# ----------------------------------------------------------------------
# Multi-instance: study-section values repeat across model sub-columns
# (T027 (b), FR-010)
# ----------------------------------------------------------------------


def test_multi_instance_article_repeats_study_section_values():
    # Two sections: one study_section ("Author"), one model_section ("Model perf").
    study_field = _field("Author", ExtractionFieldType.TEXT, parent=uuid4())
    model_field = _field("Modelling method", ExtractionFieldType.TEXT, parent=uuid4())
    study = _section("Study", ExtractionEntityRole.STUDY_SECTION, [study_field])
    model = _section("Model development", ExtractionEntityRole.MODEL_SECTION, [model_field])

    study_inst = uuid4()
    model_inst_a = uuid4()
    model_inst_b = uuid4()
    article = _article(
        "Gaca, 2011",
        study_instances={study.entity_type_id: study_inst},
        model_instances=(model_inst_a, model_inst_b),
    )
    study_fid = study.fields[0].field_id
    model_fid = model.fields[0].field_id

    data = build_workbook(
        _layout(
            sections=(study, model),
            articles=(article,),
            value_map={
                (article.run_id, study_inst, study_fid): "Gaca",
                (article.run_id, model_inst_a, model_fid): "Logistic regression",
                (article.run_id, model_inst_b, model_fid): "Cox model",
            },
        )
    )
    ws = _open(data)["CHARMS"]

    # Header row: article "Gaca, 2011" spans 2 sub-columns (merged).
    assert ws.cell(row=1, column=3).value == "Gaca, 2011"
    # The merge means cell C1 is the only one with the value; D1 is empty
    # but openpyxl's read API surfaces None for the trailing merged cell.
    assert ws.cell(row=1, column=4).value is None

    # Section "1. Study" row + field row "1.1 Author":
    # Study-section cell value MUST appear in BOTH model sub-columns
    # (the repeat-not-merge rule).
    # Find the "1.1 Author" row.
    author_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "1.1 Author":
            author_row = r
            break
    assert author_row is not None
    assert ws.cell(row=author_row, column=3).value == "Gaca"
    assert ws.cell(row=author_row, column=4).value == "Gaca"

    # Model-section field differs per sub-column.
    method_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "2.1 Modelling method":
            method_row = r
            break
    assert method_row is not None
    assert ws.cell(row=method_row, column=3).value == "Logistic regression"
    assert ws.cell(row=method_row, column=4).value == "Cox model"


# ----------------------------------------------------------------------
# Section header rows are styled bold + light fill (T027 (c), FR-009)
# ----------------------------------------------------------------------


def test_section_header_rows_have_bold_font_and_grey_fill():
    f = _field("Source", ExtractionFieldType.TEXT, parent=uuid4())
    section = _section("Source of data", ExtractionEntityRole.STUDY_SECTION, [f])
    article = _article("Gaca, 2011", study_instances={section.entity_type_id: uuid4()})
    data = build_workbook(_layout(sections=(section,), articles=(article,)))
    ws = _open(data)["CHARMS"]

    section_cell = ws.cell(row=2, column=1)
    assert section_cell.value == "1. Source of data"
    assert section_cell.font.bold is True
    fill_rgb = section_cell.fill.fgColor.rgb if section_cell.fill.fgColor else None
    assert fill_rgb is not None
    assert fill_rgb.upper().endswith("EEEEEE")


# ----------------------------------------------------------------------
# FR-019 value type formatting (T027 (d))
# ----------------------------------------------------------------------


@pytest.mark.parametrize(
    "ftype, raw_value, expected",
    [
        (ExtractionFieldType.TEXT, "Existing registry", "Existing registry"),
        (ExtractionFieldType.NUMBER, 13617, 13617),
        (ExtractionFieldType.BOOLEAN, True, "Yes"),
        (ExtractionFieldType.BOOLEAN, False, "No"),
        (ExtractionFieldType.SELECT, "Pre-operative", "Pre-operative"),
        (
            ExtractionFieldType.MULTISELECT,
            ["A", "B, with comma", "C"],
            "A; B, with comma; C",
        ),
    ],
)
def test_format_cell_per_field_type(ftype, raw_value, expected):
    f = _field("F", ftype, parent=uuid4())
    section = _section("S", ExtractionEntityRole.STUDY_SECTION, [f])
    inst = uuid4()
    article = _article("X", study_instances={section.entity_type_id: inst})
    data = build_workbook(
        _layout(
            sections=(section,),
            articles=(article,),
            value_map={(article.run_id, inst, section.fields[0].field_id): raw_value},
        )
    )
    ws = _open(data)["CHARMS"]
    assert ws.cell(row=3, column=3).value == expected


def test_none_value_renders_blank_cell():
    f = _field("F", ExtractionFieldType.TEXT, parent=uuid4())
    section = _section("S", ExtractionEntityRole.STUDY_SECTION, [f])
    article = _article("X", study_instances={section.entity_type_id: uuid4()})
    # No value in value_map → cell is blank.
    data = build_workbook(_layout(sections=(section,), articles=(article,)))
    ws = _open(data)["CHARMS"]
    assert ws.cell(row=3, column=3).value is None


# ----------------------------------------------------------------------
# Notes sheet content
# ----------------------------------------------------------------------


def test_summary_sheet_lists_omitted_articles_by_stage():
    # The omitted-by-stage tally moved from the legacy Notes sheet onto the
    # Summary sheet (the README sub-builder absorbs the rest of Notes).
    notes = ExportNotes(
        omitted_articles_by_stage={"extract": 4, "no_run": 2},
        template_version_label="CHARMS v1",
        export_mode_label="Consensus",
        generated_at=datetime(2026, 5, 23, 12, 0, 0, tzinfo=UTC),
    )
    layout = ExportLayout(
        project_name="Test Project",
        template_name="CHARMS",
        template_version=1,
        sections=(),
        articles=(),
        reviewers=(),
        mode=ExportMode.CONSENSUS,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=notes,
        value_map={},
    )
    data = build_workbook(layout)
    rows = [list(row) for row in _open(data)["Summary"].iter_rows(values_only=True)]
    flat = " ".join(str(c) for row in rows for c in row if c)
    assert "Articles omitted" in flat
    assert "stage=extract" in flat
    assert "stage=no_run" in flat


def test_ai_metadata_sheet_emits_placeholder_when_no_rows():
    data = build_workbook(_layout(include_ai_metadata=True))
    ws = _open(data)["AI metadata"]
    # Header row plus 1 placeholder row.
    assert ws.cell(row=1, column=1).value == "Article"
    assert ws.cell(row=2, column=1).value == "(No AI proposals recorded for the selected articles.)"


def test_all_users_mode_fans_out_reviewer_subcolumns():
    """FR-011 — each (article × model) splits into Consensus + N reviewers."""
    from uuid import uuid4 as _uuid4

    from app.services.extraction_export_service import (
        ExportLayout,
        ExportMode,
        ExportNotes,
        ReviewerDescriptor,
    )

    f = _field("Source", ExtractionFieldType.TEXT, parent=uuid4())
    section = _section("Source of data", ExtractionEntityRole.STUDY_SECTION, [f])
    inst_id = uuid4()
    article = _article("Gaca, 2011", study_instances={section.entity_type_id: inst_id})
    reviewer_a_id = _uuid4()
    reviewer_b_id = _uuid4()
    field_id = section.fields[0].field_id

    layout = ExportLayout(
        project_name="Test Project",
        template_name="CHARMS",
        template_version=1,
        sections=(section,),
        articles=(article,),
        reviewers=(
            ReviewerDescriptor(reviewer_id=reviewer_a_id, display_label="Alice"),
            ReviewerDescriptor(reviewer_id=reviewer_b_id, display_label="Bob"),
        ),
        mode=ExportMode.ALL_USERS,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=ExportNotes(generated_at=datetime(2026, 5, 23, tzinfo=UTC)),
        value_map={
            # consensus
            (article.run_id, inst_id, field_id, None): "Existing registry",
            # reviewer A
            (article.run_id, inst_id, field_id, reviewer_a_id): "Existing registry",
            # reviewer B (disagrees)
            (article.run_id, inst_id, field_id, reviewer_b_id): "RCT",
        },
    )
    data = build_workbook(layout)
    ws = _open(data)["CHARMS"]

    # Row 1 article header spans 3 sub-columns (Consensus + Alice + Bob).
    assert ws.cell(row=1, column=3).value == "Gaca, 2011"
    # Row 2 reviewer labels.
    assert ws.cell(row=2, column=3).value == "Consensus"
    assert ws.cell(row=2, column=4).value == "Alice"
    assert ws.cell(row=2, column=5).value == "Bob"
    # Field row (offset by reviewer header → row 4).
    # Find the row with field label
    field_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "1.1 Source":
            field_row = r
            break
    assert field_row is not None
    assert ws.cell(row=field_row, column=3).value == "Existing registry"
    assert ws.cell(row=field_row, column=4).value == "Existing registry"
    assert ws.cell(row=field_row, column=5).value == "RCT"


def test_ai_metadata_sheet_writes_proposal_rows_in_canonical_order():
    """One row per AI proposal, columns in FR-037 order."""
    from app.services.extraction_export_service import AIProposalRow, ExportLayout

    proposal = AIProposalRow(
        article_label="Gaca, 2011",
        section_label="1. Source of data",
        instance_index=1,
        field_label="1.1 Source of data",
        ai_proposed_value="Existing registry",
        confidence=0.93,
        rationale="LLM reasoning here",
        evidence_text="Patients were enrolled from the registry.",
        evidence_pages="4",
        proposed_at=datetime(2026, 5, 23, 10, 0, 0, tzinfo=UTC),
        model_used="gpt-4o-mini",
        reviewer_outcome="accepted",
        final_value_used="Existing registry",
    )
    base = _layout(include_ai_metadata=True)
    layout = ExportLayout(
        project_name=base.project_name,
        template_name=base.template_name,
        template_version=base.template_version,
        sections=base.sections,
        articles=base.articles,
        reviewers=base.reviewers,
        mode=base.mode,
        include_ai_metadata=True,
        anonymize_reviewer_names=base.anonymize_reviewer_names,
        notes=base.notes,
        value_map=base.value_map,
        ai_proposal_rows=(proposal,),
    )
    data = build_workbook(layout)
    ws = _open(data)["AI metadata"]
    # Row 2: the proposal. Columns in FR-037 order.
    assert ws.cell(row=2, column=1).value == "Gaca, 2011"
    assert ws.cell(row=2, column=2).value == "1. Source of data"
    assert ws.cell(row=2, column=3).value == 1
    assert ws.cell(row=2, column=4).value == "1.1 Source of data"
    assert ws.cell(row=2, column=5).value == "Existing registry"
    assert ws.cell(row=2, column=6).value == 0.93
    assert ws.cell(row=2, column=7).value == "LLM reasoning here"
    assert ws.cell(row=2, column=8).value == "Patients were enrolled from the registry."
    assert ws.cell(row=2, column=9).value == "4"
    # column 10: timestamp; column 11: model_used (NEW); column 12: reviewer outcome (shifted)
    assert ws.cell(row=2, column=10).value is not None
    assert ws.cell(row=2, column=11).value == "gpt-4o-mini"
    assert ws.cell(row=2, column=12).value == "accepted"
    assert ws.cell(row=2, column=13).value == "Existing registry"


def test_ai_metadata_value_columns_render_via_shared_helper():
    """The 'AI proposed value' (E) and 'Final value used' (L) columns
    pass already-resolved scalars through the shared format helper, not a
    dict-stringify path — number+unit / Yes survive intact."""
    from app.services.extraction_export_service import AIProposalRow, ExportLayout

    proposal = AIProposalRow(
        article_label="Gaca, 2011",
        section_label="1. Source of data",
        instance_index=1,
        field_label="1.1 Dose",
        ai_proposed_value="5 mg",
        confidence=0.8,
        rationale="reason",
        evidence_text="evidence",
        evidence_pages="2",
        proposed_at=datetime(2026, 6, 14, 10, 0, 0, tzinfo=UTC),
        model_used="gpt-4o",
        reviewer_outcome="accepted",
        final_value_used="Yes",
    )
    base = _layout(include_ai_metadata=True)
    layout = ExportLayout(
        project_name=base.project_name,
        template_name=base.template_name,
        template_version=base.template_version,
        sections=base.sections,
        articles=base.articles,
        reviewers=base.reviewers,
        mode=base.mode,
        include_ai_metadata=True,
        anonymize_reviewer_names=base.anonymize_reviewer_names,
        notes=base.notes,
        value_map=base.value_map,
        ai_proposal_rows=(proposal,),
    )
    ws = _open(build_workbook(layout))["AI metadata"]
    # E2 = "AI proposed value" (col 5), M2 = "Final value used" (col 13, shifted +1).
    assert ws.cell(row=2, column=5).value == "5 mg"
    assert ws.cell(row=2, column=13).value == "Yes"


def test_xlsx_safe_raises_on_dict() -> None:
    """A dict reaching _xlsx_safe means resolve_value was bypassed — it
    must fail loud, not silently str() into the sheet."""
    from app.services.exports.extraction.matrix import _xlsx_safe

    with pytest.raises(TypeError):
        _xlsx_safe({"value": 5, "unit": "mg"})


def test_workbook_emits_sheets_in_section4_order():
    """README → Summary → matrix → tidy tables → Data dictionary → Dropdown lists."""
    from app.models.extraction import (
        ExtractionCardinality,
        ExtractionEntityRole,
        ExtractionFieldType,
    )
    from app.services.exports.extraction.workbook import build_workbook
    from app.services.extraction_export_service import (
        AllowedValue,
        ArticleDescriptor,
        ExportLayout,
        ExportMode,
        ExportNotes,
        FieldDescriptor,
        FieldDictEntry,
        FrontMatter,
        SectionDescriptor,
        TidyRow,
        TidyTable,
    )

    eid = uuid4()
    fid = uuid4()
    section = SectionDescriptor(
        entity_type_id=eid,
        label="Study",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(
            FieldDescriptor(
                field_id=fid,
                label="Design",
                type=ExtractionFieldType.SELECT,
                allowed_values=("Cohort", "RCT"),
                parent_section_id=eid,
            ),
        ),
        cardinality=ExtractionCardinality.ONE,
        sort_order=0,
    )
    inst = uuid4()
    run = uuid4()
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="Gaca, 2011",
        run_id=run,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={eid: (inst,)},
    )
    fm = FrontMatter(
        project_name="P",
        template_name="CHARMS",
        template_version=1,
        export_mode_label="Consensus",
        generated_at=datetime(2026, 6, 14, tzinfo=UTC),
        article_count=1,
        record_count=1,
        contents=("README", "Summary"),
        legend=(),
        caveats=(),
        obsolete_fields_per_article={},
    )
    dict_entry = FieldDictEntry(
        field_id=fid,
        section_label="Study",
        label="Design",
        type=ExtractionFieldType.SELECT,
        unit=None,
        description=None,
        allowed_values=(AllowedValue(value="Cohort", label="Cohort"),),
        is_required=False,
        allow_other=False,
    )
    tidy = TidyTable(
        section_id=eid,
        title="Study characteristics",
        cardinality=ExtractionCardinality.ONE,
        column_field_ids=(fid,),
        column_labels=("Design",),
        rows=(
            TidyRow(
                article_id=article.article_id,
                instance_id=None,
                record_label="Gaca, 2011",
                values=("Cohort",),
            ),
        ),
    )
    layout = ExportLayout(
        project_name="P",
        template_name="CHARMS",
        template_version=1,
        sections=(section,),
        articles=(article,),
        reviewers=(),
        mode=ExportMode.CONSENSUS,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=ExportNotes(generated_at=datetime(2026, 6, 14, tzinfo=UTC)),
        value_map={(run, inst, fid): "Cohort"},
        front_matter=fm,
        data_dictionary=(dict_entry,),
        tidy_tables=(tidy,),
    )
    wb = load_workbook(io.BytesIO(build_workbook(layout)))
    assert wb.sheetnames == [
        "README",
        "Summary",
        "CHARMS",
        "Study characteristics",
        "Data dictionary",
        "Dropdown lists",
    ]


def test_workbook_emits_appraisal_sheet_after_tidy_tables() -> None:
    """Appraisal sheet appears (k+1) only when layout.appraisal is set."""
    from app.services.exports.extraction.workbook import build_workbook
    from app.services.extraction_export_service import AppraisalModel, AppraisalRow

    # Build the smallest QA layout that yields one appraisal row.
    appraisal = AppraisalModel(
        domain_section_ids=(uuid4(),),
        domain_labels=("Participants",),
        rows=(
            AppraisalRow(
                article_id=uuid4(),
                record_label="Gaca, 2011",
                domain_verdicts=("High",),
                overall="High",
                per_reviewer_overall={},
            ),
        ),
    )
    layout_with = _layout(appraisal=appraisal)
    wb = load_workbook(io.BytesIO(build_workbook(layout_with)))
    assert "Appraisal summary" in wb.sheetnames

    layout_without = _layout(appraisal=None)
    wb2 = load_workbook(io.BytesIO(build_workbook(layout_without)))
    assert "Appraisal summary" not in wb2.sheetnames

"""Unit tests for the extraction matrix builder fan-out (cardinality, any role).

The lower block (``build_matrix`` / SheetSpec assertions) is the §5.4
sub-builder lift: ``_write_main_sheet`` re-expressed as a pure ``SheetSpec``
producer, asserted on the rendered worksheet so the split keeps cell values,
merged record headers and study-section repeat-not-merge identical.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from uuid import UUID, uuid4

import pytest
from openpyxl import Workbook, load_workbook

from app.core.error_handler import AppError
from app.models.extraction import (
    ExtractionCardinality,
    ExtractionEntityRole,
    ExtractionFieldType,
)
from app.services.exports.extraction.matrix import build_matrix
from app.services.exports.extraction.sheet_spec import _render_sheet_spec
from app.services.exports.extraction.workbook import build_workbook
from app.services.extraction_export_service import (
    ArticleDescriptor,
    ExportLayout,
    ExportMode,
    ExportNotes,
    FieldDescriptor,
    SectionDescriptor,
)


def _layout(sections, articles, value_map):
    return ExportLayout(
        project_name="P",
        template_name="CHARMS",
        template_version=1,
        sections=sections,
        articles=articles,
        reviewers=(),
        mode=ExportMode.CONSENSUS,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=ExportNotes(generated_at=datetime(2026, 6, 14, tzinfo=UTC)),
        value_map=value_map,
    )


def test_many_cardinality_study_section_fans_out_one_subcolumn_per_instance():
    eid = uuid4()
    fid = uuid4()
    field = FieldDescriptor(
        field_id=fid,
        label="Index test name",
        type=ExtractionFieldType.TEXT,
        allowed_values=(),
        parent_section_id=eid,
    )
    section = SectionDescriptor(
        entity_type_id=eid,
        label="Index tests",
        role=ExtractionEntityRole.STUDY_SECTION,  # NOT a model section
        parent_entity_type_id=None,
        fields=(field,),
        cardinality=ExtractionCardinality.MANY,
    )
    inst_a, inst_b = uuid4(), uuid4()
    run_id = uuid4()
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="Gaca, 2011",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={eid: (inst_a, inst_b)},
    )
    data = build_workbook(
        _layout(
            (section,),
            (article,),
            {
                (run_id, inst_a, fid): "CT angiography",
                (run_id, inst_b, fid): "MRI",
            },
        )
    )
    ws = load_workbook(io.BytesIO(data))["CHARMS"]

    # Article header spans TWO instance sub-columns (merged C1:D1).
    assert ws.cell(row=1, column=3).value == "Gaca, 2011"
    assert ws.cell(row=1, column=4).value is None  # trailing merged cell

    field_row = None
    for r in range(2, ws.max_row + 1):
        # Hierarchical numbering prefixes the section.field index (§9).
        if ws.cell(row=r, column=2).value == "1.1 Index test name":
            field_row = r
            break
    assert field_row is not None
    # One distinct value per instance — NO collapse.
    assert ws.cell(row=field_row, column=3).value == "CT angiography"
    assert ws.cell(row=field_row, column=4).value == "MRI"


def test_column_guard_rejects_layouts_over_excel_limit():
    # One study-many section with an absurd instance count to blow the
    # 16,384-column budget deterministically (no real fan-out needed).
    eid = uuid4()
    fid = uuid4()
    field = FieldDescriptor(
        field_id=fid,
        label="F",
        type=ExtractionFieldType.TEXT,
        allowed_values=(),
        parent_section_id=eid,
    )
    section = SectionDescriptor(
        entity_type_id=eid,
        label="S",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
        cardinality=ExtractionCardinality.MANY,
    )
    run_id = uuid4()
    instance_ids = tuple(uuid4() for _ in range(16_400))
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="Big",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={eid: instance_ids},
    )
    with pytest.raises(AppError) as exc:
        build_workbook(_layout((section,), (article,), {}))
    assert "16384" in str(exc.value) or "column" in str(exc.value).lower()


# ----------------------------------------------------------------------
# §5.4 sub-builder: build_matrix -> SheetSpec, asserted on the rendered
# worksheet so the verbatim lift keeps cell values / merges identical.
# ----------------------------------------------------------------------


def _render_matrix(layout: ExportLayout):
    spec = build_matrix(layout)
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="tmp")
    _render_sheet_spec(ws, spec)
    return ws


def _spec_field(label: str, ftype: ExtractionFieldType, parent: UUID) -> FieldDescriptor:
    return FieldDescriptor(
        field_id=uuid4(),
        label=label,
        type=ftype,
        allowed_values=(),
        parent_section_id=parent,
    )


def _spec_layout(
    *,
    sections: tuple[SectionDescriptor, ...],
    articles: tuple[ArticleDescriptor, ...],
    value_map: dict | None = None,
    mode: ExportMode = ExportMode.CONSENSUS,
    reviewers: tuple = (),
) -> ExportLayout:
    return ExportLayout(
        project_name="P",
        template_name="My Template",
        template_version=1,
        sections=sections,
        articles=articles,
        reviewers=reviewers,
        mode=mode,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=ExportNotes(),
        value_map=value_map or {},
    )


def test_header_block_labels() -> None:
    sec_id = uuid4()
    field = _spec_field("1.1 Source", ExtractionFieldType.TEXT, sec_id)
    section = SectionDescriptor(
        entity_type_id=sec_id,
        label="1. Source of data",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
    )
    run_id, inst_id, art_id = uuid4(), uuid4(), uuid4()
    article = ArticleDescriptor(
        article_id=art_id,
        header_label="Smith, 2020",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={sec_id: (inst_id,)},
    )
    layout = _spec_layout(
        sections=(section,),
        articles=(article,),
        value_map={(run_id, inst_id, field.field_id): "Registry"},
    )
    ws = _render_matrix(layout)
    assert ws.cell(row=1, column=1).value == "Section"
    assert ws.cell(row=1, column=2).value == "Field"
    assert ws.cell(row=1, column=3).value == "Smith, 2020"
    # field value rendered in column C under the article.
    assert "Registry" in {ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)}


def test_empty_articles_placeholder() -> None:
    layout = _spec_layout(sections=(), articles=())
    ws = _render_matrix(layout)
    assert ws.cell(row=1, column=1).value == "(No eligible articles for the selected mode.)"


def test_study_section_repeats_not_merges_across_models() -> None:
    # The model fan-out width is driven by a MODEL_SECTION in the layout
    # (verbatim ``_article_fanout_count`` semantics — model_instances alone
    # does NOT widen the article when no model section is present). A study
    # section's value must then repeat identically across every model
    # sub-column (FR-010 repeat-not-merge), never merged.
    study_id = uuid4()
    model_id = uuid4()
    study_field = _spec_field("Author", ExtractionFieldType.TEXT, study_id)
    model_field = _spec_field("Method", ExtractionFieldType.TEXT, model_id)
    study = SectionDescriptor(
        entity_type_id=study_id,
        label="Study",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(study_field,),
    )
    model = SectionDescriptor(
        entity_type_id=model_id,
        label="Model development",
        role=ExtractionEntityRole.MODEL_SECTION,
        parent_entity_type_id=None,
        fields=(model_field,),
    )
    run_id, study_inst = uuid4(), uuid4()
    m1, m2 = uuid4(), uuid4()
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="A",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(m1, m2),
        section_instances={study_id: (study_inst,)},
    )
    layout = _spec_layout(
        sections=(study, model),
        articles=(article,),
        value_map={(run_id, study_inst, study_field.field_id): "Doe"},
    )
    ws = _render_matrix(layout)
    # Study value repeated across BOTH model sub-columns (C and D), not merged.
    row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=3).value == "Doe")
    assert ws.cell(row=row, column=4).value == "Doe"
    merged = {str(m) for m in ws.merged_cells.ranges}
    assert f"C{row}:D{row}" not in merged


def test_matrix_structural_styling() -> None:
    sec_id = uuid4()
    f1 = _spec_field("Source", ExtractionFieldType.TEXT, sec_id)
    f2 = _spec_field("Year", ExtractionFieldType.TEXT, sec_id)
    section = SectionDescriptor(
        entity_type_id=sec_id,
        label="Source of data",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(f1, f2),
    )
    run_id, inst_id = uuid4(), uuid4()
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="A",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={sec_id: (inst_id,)},
    )
    layout = _spec_layout(sections=(section,), articles=(article,))
    spec = build_matrix(layout)

    # Freeze locks label block (A,B) + header row -> first scrollable cell C2.
    assert spec.freeze == "C2"
    assert spec.tab_color is not None

    # Header row 1 is bold.
    assert spec.rows[0][0].style is not None and spec.rows[0][0].style.bold

    # The section-band row is bold + filled across the row.
    band = next(
        row
        for row in spec.rows
        if row[0].value == "1. Source of data" or row[0].value == "Source of data"
    )
    assert band[0].style is not None and band[0].style.bold
    assert band[0].style.fill is not None

    # Hierarchical numbering on field labels: section 1 -> "1.1 Source".
    labels = {row[1].value for row in spec.rows if row[1].value is not None}
    assert "1.1 Source" in labels
    assert "1.2 Year" in labels


def test_matrix_freeze_all_users_uses_two_header_rows() -> None:
    from app.services.extraction_export_service import ReviewerDescriptor

    sec_id = uuid4()
    f1 = _spec_field("Source", ExtractionFieldType.TEXT, sec_id)
    section = SectionDescriptor(
        entity_type_id=sec_id,
        label="Source",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(f1,),
    )
    rid = uuid4()
    run_id, inst_id = uuid4(), uuid4()
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="A",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={sec_id: (inst_id,)},
    )
    layout = _spec_layout(
        sections=(section,),
        articles=(article,),
        mode=ExportMode.ALL_USERS,
        reviewers=(ReviewerDescriptor(reviewer_id=rid, display_label="R1"),),
    )
    spec = build_matrix(layout)
    assert spec.freeze == "C3"


def _value_cell(ws, field: FieldDescriptor):
    """Return the rendered cell value for ``field`` in the article column (C)."""
    label_prefix = field.label
    for r in range(1, ws.max_row + 1):
        col_b = ws.cell(row=r, column=2).value
        if col_b is not None and label_prefix in str(col_b):
            return ws.cell(row=r, column=3).value
    raise AssertionError(f"no row found for field {field.label!r}")


def _single_field_layout(field: FieldDescriptor, raw_value, *, sec_id: UUID) -> ExportLayout:
    section = SectionDescriptor(
        entity_type_id=sec_id,
        label="Section",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
    )
    run_id, inst_id = uuid4(), uuid4()
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="A",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={sec_id: (inst_id,)},
    )
    return _spec_layout(
        sections=(section,),
        articles=(article,),
        value_map={(run_id, inst_id, field.field_id): raw_value},
    )


def test_boolean_false_value_renders_no_not_yes() -> None:
    # Regression for the boolean-inversion data-corruption: the consensus
    # value_map is populated with ``resolve_value`` output, which collapses a
    # ``False`` boolean to the STRING ``"No"`` (value_envelope §6). The matrix
    # cell formatter must pass that pre-resolved ``"No"`` through unchanged. A
    # ``bool("No")``-based re-implementation is truthy and would silently emit
    # ``"Yes"`` — inverting every False boolean in every exported workbook.
    from app.services.exports.value_envelope import resolve_value

    sec_id = uuid4()
    field = _spec_field("BoolFlag", ExtractionFieldType.BOOLEAN, sec_id)

    # End-to-end through the production resolve path: False -> "No".
    resolved_false = resolve_value(False, field=field)
    assert resolved_false == "No"
    ws_false = _render_matrix(_single_field_layout(field, resolved_false, sec_id=sec_id))
    assert _value_cell(ws_false, field) == "No"

    # And True -> "Yes" (the other branch, so the test is not vacuous).
    resolved_true = resolve_value(True, field=field)
    assert resolved_true == "Yes"
    ws_true = _render_matrix(_single_field_layout(field, resolved_true, sec_id=sec_id))
    assert _value_cell(ws_true, field) == "Yes"


def test_boolean_raw_false_instance_renders_no() -> None:
    # Defensive: even a raw ``False`` bool reaching the cell (field=BOOLEAN)
    # must render ``"No"`` via format_export_scalar, never ``"Yes"``.
    sec_id = uuid4()
    field = _spec_field("BoolFlag", ExtractionFieldType.BOOLEAN, sec_id)
    ws = _render_matrix(_single_field_layout(field, False, sec_id=sec_id))
    assert _value_cell(ws, field) == "No"


def test_select_value_renders_as_string() -> None:
    sec_id = uuid4()
    field = _spec_field("Choice", ExtractionFieldType.SELECT, sec_id)
    ws = _render_matrix(_single_field_layout(field, "Cohort", sec_id=sec_id))
    assert _value_cell(ws, field) == "Cohort"


def test_multiselect_list_value_joins_with_semicolons() -> None:
    sec_id = uuid4()
    field = _spec_field("Tags", ExtractionFieldType.MULTISELECT, sec_id)
    ws = _render_matrix(_single_field_layout(field, ["age", "sex", "BMI"], sec_id=sec_id))
    assert _value_cell(ws, field) == "age; sex; BMI"


def test_matrix_module_has_no_legacy_builder_dependency() -> None:
    import ast
    import pathlib

    src = pathlib.Path("app/services/exports/extraction/matrix.py").read_text(encoding="utf-8")
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.ImportFrom) and node.module:
            imported.add(node.module)
    assert "app.services.exports.extraction_xlsx_builder" not in imported

"""SC-006 / FR-026 — extraction export idempotency / byte-determinism.

Two identical builder invocations against the same in-memory layout
(no DB, no storage) MUST produce byte-identical workbooks aside from
their embedded ``generated_at`` timestamp.

Implementation note: openpyxl stamps every saved workbook with a
fresh ``modified``/``created`` time in ``docProps/core.xml`` and a
random ``application`` revision id, so bit-by-bit equality of the
raw bytes is unattainable. We assert structural equality instead:
the unpacked ZIP entries (apart from doc-properties and the
generated_at line on the Notes sheet) are identical.
"""

from __future__ import annotations

import io
import zipfile
from datetime import UTC, datetime
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.models.extraction import (
    ExtractionCardinality,
    ExtractionEntityRole,
    ExtractionFieldType,
)

# Retargeted off the legacy ``extraction_xlsx_builder`` re-export onto the pure
# orchestrator package. ``ExportColumnLimitError`` is the package's pre-build
# column guard (it doubles as an ``AppError``/``ValueError``); the plan named it
# ``ExportTooWideError`` — reconciled here to the actual exported symbol.
from app.services.exports.extraction.workbook import (
    ExportColumnLimitError,
    build_workbook,
)
from app.services.exports.extraction_snapshot_reader import AllowedValue
from app.services.extraction_export_service import (
    ArticleDescriptor,
    ExportLayout,
    ExportMode,
    ExportNotes,
    FieldDescriptor,
    FieldDictEntry,
    FrontMatter,
    ReviewerDescriptor,
    SectionDescriptor,
    TidyRow,
    TidyTable,
)

# Hard-coded UUIDs for the fixed layout — module-level so the README/Methods
# front-matter, the data dictionary, and the tidy table all reference the SAME
# ids two builds over, keeping the new sheets byte-identical across builds.
_SECTION_ID = uuid4()
_FIELD_ID = uuid4()
_RUN_ID = uuid4()
_INST_ID = uuid4()
_ARTICLE_ID = uuid4()
_GENERATED_AT = datetime(2026, 5, 23, 12, 0, 0, tzinfo=UTC)


def _fixed_layout() -> ExportLayout:
    """A layout whose every UUID is hard-coded so two runs produce the
    same workbook bytes (mod doc-properties).

    Populates the §4 publication projections — ``front_matter`` (README /
    Methods), ``data_dictionary``, ``tidy_tables``, and an explicit ``None``
    ``appraisal`` — so the new sub-builder sheets actually render and their
    XML parts are covered by the structural-determinism assertion below.
    """
    field = FieldDescriptor(
        field_id=_FIELD_ID,
        label="1.1 Source of data",
        type=ExtractionFieldType.TEXT,
        allowed_values=(),
        parent_section_id=_SECTION_ID,
    )
    section = SectionDescriptor(
        entity_type_id=_SECTION_ID,
        label="1. Source of data",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
    )
    article = ArticleDescriptor(
        article_id=_ARTICLE_ID,
        header_label="Gaca, 2011",
        run_id=_RUN_ID,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={_SECTION_ID: (_INST_ID,)},
    )

    # README / Methods front-matter (§4 #1). Its ``generated_at`` is hard-coded,
    # so the rendered "Generated at" cell is identical across builds; the only
    # surviving per-build divergence is the openpyxl docProps timestamp.
    front_matter = FrontMatter(
        project_name="Test Project",
        template_name="CHARMS",
        template_version=1,
        export_mode_label="Consensus",
        generated_at=_GENERATED_AT,
        article_count=1,
        record_count=1,
        contents=("README", "Data dictionary", "1. Source of data"),
        legend=(("—", "field not present in this Run"),),
        caveats=("Reviewer outcomes are best-effort.",),
        obsolete_fields_per_article={},
    )
    # Data dictionary (§4 #k+2): one entry with a select option set so a
    # Dropdown lists sheet co-renders and is covered by the structural diff.
    data_dictionary = (
        FieldDictEntry(
            field_id=_FIELD_ID,
            section_label="1. Source of data",
            label="1.1 Source of data",
            type=ExtractionFieldType.SELECT,
            unit=None,
            description="Where the predictor data came from.",
            allowed_values=(
                AllowedValue(value="registry", label="Existing registry"),
                AllowedValue(value="cohort", label="New cohort"),
            ),
            is_required=True,
            allow_other=False,
        ),
    )
    # Tidy table (§5.3): one records-as-rows sheet at the section grain.
    tidy_tables = (
        TidyTable(
            section_id=_SECTION_ID,
            title="1. Source of data",
            cardinality=ExtractionCardinality.ONE,
            column_field_ids=(_FIELD_ID,),
            column_labels=("1.1 Source of data",),
            rows=(
                TidyRow(
                    article_id=_ARTICLE_ID,
                    instance_id=None,
                    record_label="Gaca, 2011",
                    values=("Existing registry",),
                ),
            ),
        ),
    )
    return ExportLayout(
        project_name="Test Project",
        template_name="CHARMS",
        template_version=1,
        sections=(section,),
        articles=(article,),
        reviewers=(),
        mode=ExportMode.CONSENSUS,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=ExportNotes(
            template_version_label="CHARMS v1",
            export_mode_label="Consensus",
            generated_at=_GENERATED_AT,
        ),
        value_map={(_RUN_ID, _INST_ID, _FIELD_ID): "Existing registry"},
        front_matter=front_matter,
        data_dictionary=data_dictionary,
        tidy_tables=tidy_tables,
        appraisal=None,
    )


def _unzip_entries(blob: bytes) -> dict[str, bytes]:
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        return {name: zf.read(name) for name in sorted(zf.namelist())}


def test_build_workbook_consensus_is_deterministic_modulo_doc_properties():
    layout = _fixed_layout()
    a = build_workbook(layout)
    b = build_workbook(layout)

    entries_a = _unzip_entries(a)
    entries_b = _unzip_entries(b)
    assert sorted(entries_a.keys()) == sorted(entries_b.keys())

    # docProps/core.xml stamps modified/created times; ignore it.
    ignored = {"docProps/core.xml", "docProps/app.xml"}
    for name in entries_a:
        if name in ignored:
            continue
        assert entries_a[name] == entries_b[name], f"divergent entry: {name}"


def test_build_workbook_ai_metadata_path_is_deterministic():
    """SC-007 budget aside, the AI metadata sheet itself is deterministic
    when the layout is fixed."""
    from app.services.extraction_export_service import AIProposalRow

    base = _fixed_layout()
    proposals = (
        AIProposalRow(
            article_label="Gaca, 2011",
            section_label="1. Source of data",
            instance_index=1,
            field_label="1.1 Source of data",
            ai_proposed_value="Existing registry",
            confidence=0.93,
            rationale="LLM reasoning",
            evidence_text="excerpt",
            evidence_pages="4",
            proposed_at=datetime(2026, 5, 23, 10, 0, 0, tzinfo=UTC),
            reviewer_outcome="accepted",
            final_value_used="Existing registry",
        ),
    )
    layout = ExportLayout(
        project_name=base.project_name,
        template_name=base.template_name,
        template_version=base.template_version,
        sections=base.sections,
        articles=base.articles,
        reviewers=base.reviewers,
        mode=base.mode,
        include_ai_metadata=True,
        anonymize_reviewer_names=base.anonymize_reviewer_names,
        notes=base.notes,
        value_map=base.value_map,
        ai_proposal_rows=proposals,
    )
    a = build_workbook(layout)
    b = build_workbook(layout)

    # Open both and compare the AI metadata sheet cell-by-cell.
    wb_a = load_workbook(io.BytesIO(a))
    wb_b = load_workbook(io.BytesIO(b))
    ws_a = wb_a["AI metadata"]
    ws_b = wb_b["AI metadata"]
    rows_a = [list(r) for r in ws_a.iter_rows(values_only=True)]
    rows_b = [list(r) for r in ws_b.iter_rows(values_only=True)]
    assert rows_a == rows_b


# ----------------------------------------------------------------------
# Regression: _load_ai_proposal_rows key-shape for ALL_USERS mode
# ----------------------------------------------------------------------


@pytest.mark.asyncio
async def test_load_ai_proposal_rows_populates_final_value_for_all_users_mode() -> None:
    """Regression: when mode=ALL_USERS, the consensus value_map uses 4-tuple
    keys (run_id, instance_id, field_id, None). _load_ai_proposal_rows must
    use the 4-tuple lookup; a 3-tuple would silently produce final_value=None.
    """
    from app.services.extraction_export_service import ExtractionExportService

    run_id = uuid4()
    inst_id = uuid4()
    field_id = uuid4()
    entity_type_id = uuid4()
    article_id = uuid4()
    pid = uuid4()
    ts = datetime(2026, 5, 23, 10, 0, 0, tzinfo=UTC)

    field = FieldDescriptor(
        field_id=field_id,
        label="1.1 Source",
        type=ExtractionFieldType.TEXT,
        allowed_values=(),
        parent_section_id=entity_type_id,
    )
    section = SectionDescriptor(
        entity_type_id=entity_type_id,
        label="1. Source of data",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
    )
    article = ArticleDescriptor(
        article_id=article_id,
        header_label="Gaca, 2011",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={entity_type_id: (inst_id,)},
    )

    # ALL_USERS value_map: consensus column uses (run_id, inst_id, field_id, None).
    value_map = {(run_id, inst_id, field_id, None): "Existing registry"}

    # Mock the five DB execute calls in _load_ai_proposal_rows order:
    #   1. instance query, 2. proposal query, 3. evidence query,
    #   4. decision query, 5. entity-type label query.
    def _result(rows):
        r = MagicMock()
        r.all.return_value = rows
        return r

    mock_db = AsyncMock()
    mock_db.execute = AsyncMock(
        side_effect=[
            _result([(inst_id, entity_type_id, article_id)]),  # instances
            _result(
                [
                    (
                        pid,
                        run_id,
                        inst_id,
                        field_id,
                        {"value": "Existing registry"},
                        0.9,
                        "rationale",
                        ts,
                    )
                ]
            ),  # proposals
            _result([]),  # evidence
            _result(
                [(run_id, inst_id, field_id, uuid4(), "accept_proposal", pid)]
            ),  # decisions (reviewer-tagged)
            _result([(entity_type_id, "1. Source of data")]),  # entity labels
        ]
    )

    service = ExtractionExportService(db=mock_db, user_id="user-1", storage=MagicMock())
    rows = await service._load_ai_proposal_rows(
        articles=(article,),
        sections=(section,),
        value_map=value_map,
        mode=ExportMode.ALL_USERS,
        target_reviewer_id=None,
    )

    assert len(rows) == 1
    # Before the fix this was None; after the fix it resolves via the 4-tuple key.
    assert rows[0].final_value_used == "Existing registry"


# ----------------------------------------------------------------------
# S5 split: structural determinism of the restyled multi-sheet workbook
# + the exact 16,384-column guard boundary (via _matrix_column_count).
# ----------------------------------------------------------------------


def test_styled_matrix_is_structurally_deterministic() -> None:
    layout = _fixed_layout()
    a = build_workbook(layout)
    b = build_workbook(layout)

    def _entries(data: bytes) -> dict[str, bytes]:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return {n: zf.read(n) for n in zf.namelist() if not n.startswith("docProps/")}

    ea, eb = _entries(a), _entries(b)
    assert ea.keys() == eb.keys()
    for name in ea:
        # The Notes sheet carries generated_at; every other part (incl. the
        # styled matrix sheet + styles.xml) must be byte-identical.
        if "notes" in name.lower() or name.endswith("sharedStrings.xml"):
            continue
        assert ea[name] == eb[name], f"non-deterministic part: {name}"


def test_column_guard_boundary() -> None:
    from app.services.exports.extraction.workbook import _matrix_column_count

    # Build a layout whose matrix is exactly at the limit, and one over it.
    def _n_article_layout(n: int) -> ExportLayout:
        sec_id = uuid4()
        field = FieldDescriptor(
            field_id=uuid4(),
            label="F",
            type=ExtractionFieldType.TEXT,
            allowed_values=(),
            parent_section_id=sec_id,
        )
        section = SectionDescriptor(
            entity_type_id=sec_id,
            label="S",
            role=ExtractionEntityRole.STUDY_SECTION,
            parent_entity_type_id=None,
            fields=(field,),
        )
        articles = tuple(
            ArticleDescriptor(
                article_id=uuid4(),
                header_label=f"a{i}",
                run_id=uuid4(),
                run_stage=None,
                version_id=None,
                model_instances=(),
                # One cardinality=one study section ⇒ a single data column
                # per article (no model / many-axis fan-out).
                section_instances={sec_id: (uuid4(),)},
            )
            for i in range(n)
        )
        return ExportLayout(
            project_name="P",
            template_name="T",
            template_version=1,
            sections=(section,),
            articles=articles,
            reviewers=(),
            mode=ExportMode.CONSENSUS,
            include_ai_metadata=False,
            anonymize_reviewer_names=False,
            notes=ExportNotes(),
            value_map={},
        )

    at_limit = _n_article_layout(16_382)  # 2 + 16382 = 16384
    assert _matrix_column_count(at_limit) == 16_384
    build_workbook(at_limit)  # must not raise

    with pytest.raises(ValueError):
        build_workbook(_n_article_layout(16_383))


# ----------------------------------------------------------------------
# S8 / Task 73: extend structural determinism to the new publication sheets
# (README/Methods, Data dictionary, tidy tables, Dropdown lists) + a
# 500×100 all-users case that blows past the 16,384-column guard.
# ----------------------------------------------------------------------


def _structural_entries(blob: bytes) -> dict[str, bytes]:
    """Unpacked ZIP parts MINUS the inherently non-deterministic bits.

    openpyxl stamps a fresh ``modified``/``created`` time and a random
    application revision into ``docProps/*`` on every save, so those two
    parts can never be byte-stable. Everything else — including the new
    sub-builder sheet XML and the README/Methods front-matter sheet — must
    be byte-identical for a fixed layout. The README ``Generated at`` cell
    is part of that "everything else" but is pinned to ``_GENERATED_AT``, so
    it is deterministic by construction; this allow-list still excludes the
    docProps parts that carry the wall-clock save time.
    """
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        return {n: zf.read(n) for n in zf.namelist() if not n.startswith("docProps/")}


def _sheet_titles(blob: bytes) -> list[str]:
    return load_workbook(io.BytesIO(blob)).sheetnames


def test_new_sheets_are_structurally_deterministic() -> None:
    layout = _fixed_layout()
    a = _structural_entries(build_workbook(layout))
    b = _structural_entries(build_workbook(layout))
    assert a.keys() == b.keys()
    for name in a:
        assert a[name] == b[name], f"divergent entry: {name}"

    # The new sub-builder sheets are present in the rendered workbook.
    names = _sheet_titles(build_workbook(layout))
    assert "Data dictionary" in names
    assert any(n.startswith("README") or "Methods" in n for n in names)
    # The tidy table + co-located dropdown catalogue render too (the
    # _fixed_layout dictionary entry carries select options).
    assert "Dropdown lists" in names
    assert "1. Source of data" in names


def _wide_all_users_layout(*, n_articles: int, subcols_each: int) -> ExportLayout:
    """An ALL_USERS layout whose matrix fans out to
    ``n_articles × subcols_each × (1 consensus + len(reviewers))`` data
    columns — the reviewer axis is what pushes 500×100 past the guard.

    Mirrors the ``_wide_layout`` builder in
    ``test_extraction_export_column_guard.py``: a single ``cardinality=MANY``
    section so each article's ``section_instances`` fan out one data column
    per instance, here multiplied by the all-users reviewer axis.
    """
    sec_id = uuid4()
    field = FieldDescriptor(
        field_id=uuid4(),
        label="F",
        type=ExtractionFieldType.TEXT,
        allowed_values=(),
        parent_section_id=sec_id,
    )
    section = SectionDescriptor(
        entity_type_id=sec_id,
        label="Sec",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
        cardinality=ExtractionCardinality.MANY,
    )
    articles = tuple(
        ArticleDescriptor(
            article_id=uuid4(),
            header_label=f"A{i}",
            run_id=uuid4(),
            run_stage=None,
            version_id=uuid4(),
            model_instances=(),
            section_instances={sec_id: tuple(uuid4() for _ in range(subcols_each))},
        )
        for i in range(n_articles)
    )
    reviewers = (
        ReviewerDescriptor(reviewer_id=uuid4(), display_label="Reviewer A"),
        ReviewerDescriptor(reviewer_id=uuid4(), display_label="Reviewer B"),
    )
    return ExportLayout(
        project_name="P",
        template_name="T",
        template_version=1,
        sections=(section,),
        articles=articles,
        reviewers=reviewers,
        mode=ExportMode.ALL_USERS,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=ExportNotes(),
        value_map={},
    )


def test_500x100_all_users_exceeds_column_guard() -> None:
    # 500 articles × 100 reviewer/instance sub-columns each blows past 16,384.
    layout = _wide_all_users_layout(n_articles=500, subcols_each=100)
    with pytest.raises(ExportColumnLimitError):
        build_workbook(layout)


# ----------------------------------------------------------------------
# S7 / Task 63: appraisal-summary determinism + worst-case tie-break.
#
# Mirrors the A6 red-green-across-seeds discipline: prove byte-stable
# output and tie-break stability, not just one happy path. The verdict
# rollup must (a) be reproducible call-to-call, (b) resolve ties to the
# first-encountered worst label, and (c) pin reviewer Overall columns to
# layout.reviewers order regardless of the order of NON-order-bearing
# inputs (value_map insertion order, the sections tuple the model build
# sorts by sort_order). Reviewer column order is order-bearing and fixed
# by layout.reviewers — the model never re-sorts reviewers.
# ----------------------------------------------------------------------


def _appraisal_layout(
    *,
    sections,
    articles,
    reviewers,
    value_map,
) -> ExportLayout:
    """All-users ExportLayout carrying a model built by _build_appraisal_model.

    Builds the AppraisalModel through the real service helper so the
    determinism assertion covers the full model-build -> render path.
    """
    from app.services.extraction_export_service import ExtractionExportService

    appraisal = ExtractionExportService._build_appraisal_model(
        sections=sections,
        articles=articles,
        reviewers=reviewers,
        value_map=value_map,
        mode=ExportMode.ALL_USERS,
    )
    return ExportLayout(
        project_name="QA Project",
        template_name="PROBAST",
        template_version=1,
        sections=sections,
        articles=articles,
        reviewers=reviewers,
        mode=ExportMode.ALL_USERS,
        include_ai_metadata=False,
        anonymize_reviewer_names=False,
        notes=ExportNotes(
            template_version_label="PROBAST v1",
            export_mode_label="All users",
            generated_at=datetime(2026, 5, 23, 12, 0, 0, tzinfo=UTC),
        ),
        value_map=value_map,
        appraisal=appraisal,
    )


def _tied_appraisal_inputs():
    """Two domains tied at the worst rank, but with DISTINCT verdict text.

    The two worst-position verdicts are ``High`` (domain 1, first in sort_order)
    and ``high`` (domain 2) — same severity rank, different byte strings. That
    distinction is deliberate: it makes the first-wins tie-break OBSERVABLE.
    First-wins preserves the first original text ``High``; a last-wins rollup
    would surface ``high`` instead. Were both literally ``High`` (byte-identical),
    the two tied candidates would be indistinguishable and the headline tie-break
    deliverable would be untested (a `> -> >=` mutation in _appraisal_overall
    would survive). _appraisal_overall is label-preserving precisely so this
    direction is assertable.

    Two reviewers, both fed the same tied High/high pair, so the tie-break is
    exercised in the consensus column AND in every per-reviewer Overall column.

    All UUIDs are hard-coded per-call (fresh uuid4) so a single invocation is
    internally consistent; the test reuses ONE invocation's ids across reshuffles
    so only input ORDER — never identity — varies.
    """
    from app.models.extraction import (
        ExtractionEntityRole,
        ExtractionFieldType,
    )

    risk_labels = ("Low", "Unclear", "High")

    def _verdict_field(parent):
        return FieldDescriptor(
            field_id=uuid4(),
            label="Risk of bias",
            type=ExtractionFieldType.SELECT,
            allowed_values=risk_labels,
            parent_section_id=parent,
        )

    sid1, sid2 = uuid4(), uuid4()
    f1 = _verdict_field(sid1)
    f2 = _verdict_field(sid2)
    d1 = SectionDescriptor(
        entity_type_id=sid1,
        label="Participants",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(f1,),
        cardinality=ExtractionCardinality.ONE,
        sort_order=0,
    )
    d2 = SectionDescriptor(
        entity_type_id=sid2,
        label="Predictors",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(f2,),
        cardinality=ExtractionCardinality.ONE,
        sort_order=1,
    )

    run_id = uuid4()
    inst1, inst2 = uuid4(), uuid4()
    article = ArticleDescriptor(
        article_id=uuid4(),
        header_label="Gaca, 2011",
        run_id=run_id,
        run_stage=None,
        version_id=None,
        model_instances=(),
        section_instances={sid1: (inst1,), sid2: (inst2,)},
    )

    r1, r2 = uuid4(), uuid4()
    reviewers = (
        ReviewerDescriptor(reviewer_id=r1, display_label="Reviewer A"),
        ReviewerDescriptor(reviewer_id=r2, display_label="Reviewer B"),
    )

    # All-users 4-tuple keys. Domain 1 (sort_order=0, first in the rollup) carries
    # ``High``; domain 2 carries the same-rank-but-distinct ``high``. Both rank
    # equally as the worst case, so the rollup must pick the FIRST — ``High`` —
    # for the consensus column AND for each reviewer. The differing byte strings
    # give the first-wins tie-break teeth (a last-wins rollup would yield ``high``).
    value_map = {
        (run_id, inst1, f1.field_id, None): "High",
        (run_id, inst2, f2.field_id, None): "high",
        (run_id, inst1, f1.field_id, r1): "High",
        (run_id, inst2, f2.field_id, r1): "high",
        (run_id, inst1, f1.field_id, r2): "High",
        (run_id, inst2, f2.field_id, r2): "high",
    }
    return (d1, d2), (article,), reviewers, value_map


def test_appraisal_summary_spec_is_reproducible_and_tie_breaks_first_worst() -> None:
    from app.services.exports.extraction.appraisal_summary import build_appraisal_summary

    sections, articles, reviewers, value_map = _tied_appraisal_inputs()
    layout = _appraisal_layout(
        sections=sections,
        articles=articles,
        reviewers=reviewers,
        value_map=value_map,
    )

    spec_a = build_appraisal_summary(layout)
    spec_b = build_appraisal_summary(layout)
    assert spec_a is not None

    def _rows(spec):
        return tuple(tuple((c.value, c.style) for c in row) for row in spec.rows)

    # (a) Identical SheetSpec across repeated calls — header + every data row.
    assert _rows(spec_a) == _rows(spec_b)
    assert spec_a.column_widths == spec_b.column_widths
    assert spec_a.freeze == spec_b.freeze

    # (b) Tied worst case rolls up to the FIRST-encountered worst label. Domain 1
    # is ``High`` and domain 2 is the same-rank ``high``; first-wins must surface
    # ``High`` (the first), never ``high``. This is the load-bearing assertion:
    # because the two tied labels differ in text, it distinguishes first-wins from
    # last-wins (a `> -> >=` mutation in _appraisal_overall flips it to ``high``
    # and fails here). Every Overall column — consensus + both reviewers — is fed
    # the same tied pair, so all three pin to ``High``.
    header = tuple(c.value for c in spec_a.rows[0])
    assert header == (
        "Record",
        "Participants",
        "Predictors",
        "Overall",
        "Overall — Reviewer A",
        "Overall — Reviewer B",
    )
    data = tuple(c.value for c in spec_a.rows[1])
    assert data == ("Gaca, 2011", "High", "high", "High", "High", "High")


def test_appraisal_reviewer_columns_pinned_to_layout_order() -> None:
    """Reviewer Overall columns follow layout.reviewers order, not the dict
    iteration order of per_reviewer_overall. Reversing layout.reviewers must
    reverse the reviewer columns (proving the order is read from the tuple, not
    leaked from a dict)."""
    from app.services.exports.extraction.appraisal_summary import build_appraisal_summary

    sections, articles, reviewers, value_map = _tied_appraisal_inputs()

    forward = build_appraisal_summary(
        _appraisal_layout(
            sections=sections,
            articles=articles,
            reviewers=reviewers,
            value_map=value_map,
        )
    )
    reversed_reviewers = tuple(reversed(reviewers))
    backward = build_appraisal_summary(
        _appraisal_layout(
            sections=sections,
            articles=articles,
            reviewers=reversed_reviewers,
            value_map=value_map,
        )
    )

    fwd_header = tuple(c.value for c in forward.rows[0])
    bwd_header = tuple(c.value for c in backward.rows[0])
    assert fwd_header[-2:] == ("Overall — Reviewer A", "Overall — Reviewer B")
    assert bwd_header[-2:] == ("Overall — Reviewer B", "Overall — Reviewer A")


def test_appraisal_workbook_is_byte_stable_across_value_map_reshuffles() -> None:
    """The appraisal rollup must not leak value_map dict-iteration order into
    the output: reshuffling the value_map insertion order under a seeded RNG
    (a NON-order-bearing input — article iteration and reviewer columns are the
    only order-bearing axes, both held fixed) must yield a byte-identical
    workbook (mod doc-properties) every time."""
    import random

    sections, articles, reviewers, value_map = _tied_appraisal_inputs()

    baseline = build_workbook(
        _appraisal_layout(
            sections=sections,
            articles=articles,
            reviewers=reviewers,
            value_map=value_map,
        )
    )
    base_entries = _unzip_entries(baseline)
    ignored = {"docProps/core.xml", "docProps/app.xml"}

    rng = random.Random(20260615)
    for _ in range(5):
        shuffled_items = list(value_map.items())
        rng.shuffle(shuffled_items)
        shuffled_value_map = dict(shuffled_items)

        # sections order (order-bearing for the matrix) and reviewers order
        # (order-bearing for the reviewer columns) are held FIXED; only the
        # value_map dict insertion order varies.
        layout = _appraisal_layout(
            sections=sections,
            articles=articles,
            reviewers=reviewers,
            value_map=shuffled_value_map,
        )
        blob = build_workbook(layout)
        entries = _unzip_entries(blob)
        assert entries.keys() == base_entries.keys()
        for name in base_entries:
            if name in ignored:
                continue
            assert entries[name] == base_entries[name], f"divergent entry: {name}"

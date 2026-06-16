"""Orchestrator: spec-order sheet assembly + the §5.5 column guard."""

from __future__ import annotations

import io
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.models.extraction import ExtractionEntityRole, ExtractionFieldType
from app.services.exports.extraction.workbook import build_workbook
from app.services.extraction_export_service import (
    ArticleDescriptor,
    ExportLayout,
    ExportMode,
    ExportNotes,
    FieldDescriptor,
    SectionDescriptor,
)


def _one_field_layout(*, include_ai: bool, n_articles: int = 1) -> ExportLayout:
    sec_id = uuid4()
    field = FieldDescriptor(
        field_id=uuid4(),
        label="Source",
        type=ExtractionFieldType.TEXT,
        allowed_values=(),
        parent_section_id=sec_id,
    )
    section = SectionDescriptor(
        entity_type_id=sec_id,
        label="Source of data",
        role=ExtractionEntityRole.STUDY_SECTION,
        parent_entity_type_id=None,
        fields=(field,),
    )
    articles = tuple(
        ArticleDescriptor(
            article_id=uuid4(),
            header_label=f"Art {i}",
            run_id=uuid4(),
            run_stage=None,
            version_id=None,
            model_instances=(),
            section_instances={sec_id: (uuid4(),)},
        )
        for i in range(n_articles)
    )
    return ExportLayout(
        project_name="P",
        template_name="My Template",
        template_version=1,
        sections=(section,),
        articles=articles,
        reviewers=(),
        mode=ExportMode.CONSENSUS,
        include_ai_metadata=include_ai,
        anonymize_reviewer_names=False,
        notes=ExportNotes(),
        value_map={},
    )


def test_sheet_order_without_ai() -> None:
    # §4 order: README (absorbs Notes) → Summary → matrix → Data dictionary.
    # This layout populates neither tidy_tables nor a non-empty data
    # dictionary, so no tidy / Dropdown lists sheets are emitted.
    data = build_workbook(_one_field_layout(include_ai=False))
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["README", "Summary", "My Template", "Data dictionary"]


def test_sheet_order_with_ai() -> None:
    # AI metadata is the trailing optional sheet, after the §4 specs.
    data = build_workbook(_one_field_layout(include_ai=True))
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == [
        "README",
        "Summary",
        "My Template",
        "Data dictionary",
        "AI metadata",
    ]


def test_column_guard_rejects_oversized_layout() -> None:
    # 16,385 articles × 1 single-instance column each blows past Excel's
    # 16,384-column ceiling once the 2 label cols are added.
    layout = _one_field_layout(include_ai=False, n_articles=16_385)
    with pytest.raises(ValueError, match="16384|column"):
        build_workbook(layout)
